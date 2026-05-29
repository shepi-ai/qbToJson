#!/usr/bin/env python3
"""
Customer Concentration (Sales by Customer Summary) Converter
Converts CSV, XLSX, and PDF Sales by Customer Summary reports to simplified JSON array format
"""

import csv
import re
import sys
from pathlib import Path
import argparse
from typing import Dict, List, Any, Optional

from base_converter import BaseConverter

# Import guards from base (openpyxl, pdfplumber)
from base_converter import XLSX_SUPPORT, PDF_SUPPORT
if XLSX_SUPPORT:
    import openpyxl
if PDF_SUPPORT:
    import pdfplumber


class CustomerConcentrationConverter(BaseConverter):
    """Converts Sales by Customer Summary reports to simplified JSON array format"""

    def __init__(self, **kwargs):
        # Concentration converters don't use account lookup
        super().__init__(use_account_lookup=False)

    def calculate_percentages(self, customers: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Calculate percentage of total for each customer"""
        grand_total = sum(c['revenue'] for c in customers)

        for customer in customers:
            if grand_total > 0:
                customer['percentage'] = (customer['revenue'] / grand_total) * 100
            else:
                customer['percentage'] = 0.0

        customers.sort(key=lambda x: x['revenue'], reverse=True)
        return customers

    def parse_csv(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse CSV file and convert to simplified JSON array format"""
        customer_map = {}

        with open(filepath, 'r', encoding='utf-8') as f:
            # Skip header lines
            for _ in range(4):
                f.readline()

            reader = csv.DictReader(f)
            current_parent = None

            for row in reader:
                customer_name = row.get('Customer', '').strip()

                # Skip empty rows or TOTAL row
                if not customer_name or customer_name.upper() == 'TOTAL':
                    break

                # Check if this is a "Total for" row (sub-customer total)
                if customer_name.startswith('Total for '):
                    parent_name = customer_name.replace('Total for ', '')
                    if parent_name in customer_map:
                        total = self.parse_amount(row.get('Total', '0'))
                        customer_map[parent_name]['revenue'] = total
                    current_parent = None
                    continue

                total = self.parse_amount(row.get('Total', '0'))

                if total == 0.0:
                    # Could be a parent (group header) or standalone zero-revenue customer
                    # Add as entry either way; "Total for" rows will update parents
                    current_parent = customer_name
                    if customer_name not in customer_map:
                        customer_map[customer_name] = {
                            'customerName': customer_name,
                            'revenue': 0.0,
                            'percentage': 0.0
                        }
                    continue

                if current_parent and current_parent in customer_map:
                    customer_map[current_parent]['revenue'] += total
                else:
                    if customer_name not in customer_map:
                        customer_map[customer_name] = {
                            'customerName': customer_name,
                            'revenue': total,
                            'percentage': 0.0
                        }

        return self.calculate_percentages(list(customer_map.values()))

    def _parse_pivot_table(self, sheet) -> List[Dict[str, Any]]:
        """Parse P&L by Customer pivot table format"""
        customer_map = {}
        
        # Find the customer header row (contains all customer names)
        customer_row_idx = None
        customer_names = []
        
        for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if not row or not row[0]:
                continue
            # Look for row with many non-empty cells (customer names)
            non_empty = sum(1 for cell in row[1:] if cell and str(cell).strip())
            if non_empty > 20:  # Pivot table has 30+ customer columns (Q1 reports have fewer)
                customer_row_idx = idx
                customer_names = [str(cell).strip() if cell else None for cell in row]
                print(f"[CUSTOMER-CONC] Found customer header row at {idx} with {non_empty} customers")
                break
        
        if not customer_row_idx:
            raise ValueError("Could not find customer header row in pivot table")
        
        # Find "Total Income" row or similar
        total_income_row = None
        for idx, row in enumerate(sheet.iter_rows(min_row=customer_row_idx + 1, values_only=True), customer_row_idx + 1):
            if not row or not row[0]:
                continue
            first_cell = str(row[0]).strip().lower()
            if 'total income' in first_cell or first_cell == 'income':
                total_income_row = idx
                print(f"[CUSTOMER-CONC] Found Total Income row at {idx}")
                break
        
        if not total_income_row:
            raise ValueError("Could not find Total Income row in pivot table")
        
        # Extract values from Total Income row
        income_row = list(sheet.iter_rows(min_row=total_income_row, max_row=total_income_row, values_only=True))[0]
        
        # Map customer names to their sales values
        for col_idx, customer_name in enumerate(customer_names):
            if not customer_name or customer_name.upper() in ['TOTAL', 'NOT SPECIFIED']:
                continue
            
            if col_idx < len(income_row):
                value = income_row[col_idx]
                if value is not None:
                    revenue = self.parse_amount(str(value))
                    if revenue > 0:  # Only include customers with sales
                        customer_map[customer_name] = {
                            'customerName': customer_name,
                            'revenue': revenue,
                            'percentage': 0.0
                        }
        
        print(f"[CUSTOMER-CONC] Extracted {len(customer_map)} customers with sales from pivot table")
        return self.calculate_percentages(list(customer_map.values()))

    def _sum_row_values(self, row, value_col_indices: List[int]) -> float:
        """Sum all numeric values across the given column indices for a row"""
        total = 0.0
        for idx in value_col_indices:
            if idx < len(row) and row[idx] is not None:
                total += self.parse_amount(str(row[idx]))
        return total

    def parse_xlsx(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse XLSX file and convert to simplified JSON array format"""
        self.check_xlsx_support()

        customer_map = {}

        workbook = openpyxl.load_workbook(filepath, data_only=True)
        sheet = workbook.active

        # Find header row - look for row where first cell is exactly "Customer"
        header_row = None
        for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row and row[0] and str(row[0]).strip() == 'Customer':
                header_row = idx
                break

        if not header_row:
            # Fallback: look for any row where a cell is exactly 'Customer'
            for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if row and any(str(cell).strip() == 'Customer' for cell in row if cell):
                    header_row = idx
                    break

        # A standard Sales by Customer Summary lists customers down rows under a
        # "Customer" header; a wide column count just means the report spans many
        # months. Only when there is NO "Customer" header row do we treat the sheet
        # as a P&L by Customer pivot, where customers run across the columns.
        if not header_row:
            if sheet.max_column > 30:
                print(f"[CUSTOMER-CONC] No 'Customer' header row; treating as pivot table ({sheet.max_column} columns)")
                return self._parse_pivot_table(sheet)
            raise ValueError("Could not find header row in XLSX file")

        headers = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        col_map = {str(header).strip(): idx for idx, header in enumerate(headers) if header}

        # Determine value columns: prefer 'Total' column, otherwise sum all non-Customer columns
        has_total_col = 'Total' in col_map
        customer_col = col_map.get('Customer', 0)
        if has_total_col:
            value_col_indices = [col_map['Total']]
        else:
            # Monthly breakdown - sum all numeric columns except Customer
            value_col_indices = [idx for idx, header in enumerate(headers) if header and idx != customer_col]

        current_parent = None

        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[customer_col]:
                continue

            customer_name = str(row[customer_col]).strip()

            if customer_name.upper() == 'CUSTOMER':
                continue

            if 'Sandbox Company' in customer_name or 'Sales by Customer' in customer_name:
                continue

            # Skip rows that are month-column headers (e.g. "May" or "May 2024"),
            # but NOT real customers whose name merely contains a month substring
            # (e.g. "Mayert-Lehner", "Marchetti").
            if re.fullmatch(
                r'(January|February|March|April|May|June|July|August|September|October|November|December)(\s+\d{4})?',
                customer_name,
                re.IGNORECASE,
            ):
                continue

            if customer_name.upper() == 'TOTAL':
                break

            if customer_name.startswith('Total for '):
                parent_name = customer_name.replace('Total for ', '')
                if parent_name in customer_map:
                    total = self._sum_row_values(row, value_col_indices)
                    customer_map[parent_name]['revenue'] = total
                current_parent = None
                continue

            total = self._sum_row_values(row, value_col_indices)

            if total == 0.0:
                # Check if this is a parent by looking ahead (has "Total for" row)
                # For now, add it as a potential parent AND as an entry
                current_parent = customer_name
                if customer_name not in customer_map:
                    customer_map[customer_name] = {
                        'customerName': customer_name,
                        'revenue': 0.0,
                        'percentage': 0.0
                    }
                continue

            if current_parent and current_parent in customer_map:
                customer_map[current_parent]['revenue'] += total
            else:
                if customer_name not in customer_map:
                    customer_map[customer_name] = {
                        'customerName': customer_name,
                        'revenue': total,
                        'percentage': 0.0
                    }

        return self.calculate_percentages(list(customer_map.values()))

    # Table extraction settings: QB PDF reports draw no cell borders, so rely on
    # text alignment to recover columns.
    _PDF_TABLE_SETTINGS = {"vertical_strategy": "text", "horizontal_strategy": "text"}

    @staticmethod
    def _row_has_amount(cells: List[str]) -> bool:
        """True if any non-name cell contains a digit (a monetary value)."""
        return any(re.search(r'\d', c) for c in cells[1:])

    def _row_total(self, cells: List[str]) -> float:
        """Grand-total value for a row = the last (TOTAL) column."""
        last = cells[-1]
        return self.parse_amount(last) if last and last != '-' else 0.0

    def _gather_pdf_total_rows(self, filepath: Path) -> List[List[str]]:
        """Collect data rows from the report's grand-total section.

        Works for both the single-column "summary" PDF and the multi-page
        "monthly" PDF: the latter repeats every customer once per month-group,
        but only the final group's header ends in a TOTAL column, so we keep
        rows only while that grand-total column is active.
        """
        rows: List[List[str]] = []
        active = False
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                for raw in (page.extract_table(self._PDF_TABLE_SETTINGS) or []):
                    cells = [(c or "").strip() for c in raw]
                    if not any(cells):
                        continue
                    name = cells[0]
                    if name.upper().startswith("CUSTOMER"):
                        non_empty = [c for c in cells if c]
                        active = bool(non_empty) and non_empty[-1].upper() == "TOTAL"
                        continue
                    if not active or not name:
                        continue
                    # Skip report furniture: footers and the "2023 2024 ..." year sub-header.
                    if 'Accrual Basis' in name or 'GMTZ' in name or re.fullmatch(r'(\d{4} ?)+', name):
                        continue
                    rows.append(cells)
        return rows

    def parse_pdf(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse PDF file and convert to simplified JSON array format.

        QB wraps long customer names and "Total for" labels onto a second line,
        which the table extractor returns as a separate amount-less row. We
        disambiguate such rows: a parent header is the only one followed by a
        "Total for" line (its children's amounts roll up to it); any other
        amount-less row is the wrapped tail of the row above it.
        """
        self.check_pdf_support()

        rows = self._gather_pdf_total_rows(filepath)
        customer_map: Dict[str, Dict[str, Any]] = {}
        parent = None        # current parent customer accumulating sub-customer sales
        last_key = None       # most recent standalone customer (for name-wrap continuation)
        after_total_for = False

        def looks_like_parent(i: int) -> bool:
            # A parent is followed by its children (amount rows) and then a
            # "Total for" line, before any other amount-less row appears.
            for j in range(i + 1, len(rows)):
                if rows[j][0].startswith("Total for"):
                    return True
                if not self._row_has_amount(rows[j]):
                    return False
            return False

        for i, cells in enumerate(rows):
            name = cells[0]

            if name.startswith("Total for"):
                if parent and parent in customer_map:
                    customer_map[parent]['revenue'] = self._row_total(cells)
                parent = None
                last_key = None
                after_total_for = True
                continue

            if name.upper() == "TOTAL":      # grand total line
                after_total_for = False
                continue

            if self._row_has_amount(cells):
                total = self._row_total(cells)
                if parent and parent in customer_map:
                    customer_map[parent]['revenue'] += total
                    last_key = None
                else:
                    customer_map[name] = {'customerName': name, 'revenue': total, 'percentage': 0.0}
                    last_key = name
                after_total_for = False
                continue

            # Amount-less row: parent header, label wrap, or name wrap.
            if looks_like_parent(i):
                parent = name
                customer_map.setdefault(name, {'customerName': name, 'revenue': 0.0, 'percentage': 0.0})
                last_key = name
            elif after_total_for:
                pass                          # wrapped tail of the "Total for X" label
            elif last_key and last_key in customer_map:
                merged = f"{last_key} {name}".strip()   # wrapped tail of previous customer name
                entry = customer_map.pop(last_key)
                entry['customerName'] = merged
                customer_map[merged] = entry
                last_key = merged
            after_total_for = False

        print(f"[CUSTOMER-CONC-PARSER] Parsed {len(customer_map)} customers")
        return self.calculate_percentages(list(customer_map.values()))


def main():
    parser = argparse.ArgumentParser(description='Convert Sales by Customer Summary reports to JSON format')
    parser.add_argument('input', help='Input file (CSV, XLSX, or PDF)')
    parser.add_argument('-o', '--output', help='Output JSON file (default: print to stdout)')
    parser.add_argument('--batch', action='store_true', help='Process all files in a directory')

    args = parser.parse_args()

    converter = CustomerConcentrationConverter()

    if args.batch:
        input_path = Path(args.input)
        if not input_path.is_dir():
            print(f"Error: {input_path} is not a directory", file=sys.stderr)
            sys.exit(1)

        output_dir = Path(args.output) if args.output else input_path.parent / 'converted'
        output_dir.mkdir(exist_ok=True)

        for file in input_path.glob('*'):
            if file.suffix.lower() in ['.csv', '.xlsx', '.pdf']:
                try:
                    output_file = output_dir / f"{file.stem}_customer_concentration.json"
                    result = converter.convert_to_json(file, output_file)
                    print(result)
                except Exception as e:
                    print(f"Error processing {file}: {e}", file=sys.stderr)
    else:
        input_path = Path(args.input)
        if not input_path.exists():
            print(f"Error: {input_path} does not exist", file=sys.stderr)
            sys.exit(1)

        try:
            if args.output:
                result = converter.convert_to_json(input_path, Path(args.output))
                print(result)
            else:
                print(converter.convert_to_json(input_path))
        except Exception as e:
            print(f"Error: {e}", file=sys.stderr)
            sys.exit(1)


if __name__ == '__main__':
    main()
