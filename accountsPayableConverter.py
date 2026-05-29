#!/usr/bin/env python3
"""
Accounts Payable (A/P Aging Summary) Converter
Converts CSV, XLSX, and PDF A/P Aging Summary reports to QuickBooks JSON format
"""

import csv
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
import argparse
from typing import Dict, List, Any, Optional

from base_converter import BaseConverter, XLSX_SUPPORT, PDF_SUPPORT

if XLSX_SUPPORT:
    import openpyxl
if PDF_SUPPORT:
    import pdfplumber


class AccountsPayableConverter(BaseConverter):
    """Converts A/P Aging Summary reports to QuickBooks-style JSON format"""

    def __init__(self, **kwargs):
        super().__init__(use_account_lookup=False)
        self.report_date = datetime.now(timezone.utc).strftime('%Y-%m-%d')

    def create_header(self, report_date: str = None) -> Dict[str, Any]:
        """Create QuickBooks-style header for AgedPayables report"""
        if report_date is None:
            report_date = self.report_date
        return self.make_qb_header(
            "AgedPayables",
            start_period=report_date,
            end_period=report_date,
            options=[
                {"name": "report_date", "value": report_date},
                {"name": "NoReportData", "value": "false"}
            ]
        )

    def create_columns(self) -> Dict[str, List[Dict[str, Any]]]:
        """Create column definitions for A/P Aging report"""
        return {
            "column": [
                {"colTitle": "", "colType": "Vendor", "metaData": [], "columns": None},
                {"colTitle": "Current", "colType": "Money", "metaData": [{"name": "ColKey", "value": "current"}], "columns": None},
                {"colTitle": "1 - 30", "colType": "Money", "metaData": [{"name": "ColKey", "value": "0"}], "columns": None},
                {"colTitle": "31 - 60", "colType": "Money", "metaData": [{"name": "ColKey", "value": "1"}], "columns": None},
                {"colTitle": "61 - 90", "colType": "Money", "metaData": [{"name": "ColKey", "value": "2"}], "columns": None},
                {"colTitle": "91 and over", "colType": "Money", "metaData": [{"name": "ColKey", "value": "3"}], "columns": None},
                {"colTitle": "Total", "colType": "Money", "metaData": [{"name": "ColKey", "value": "total"}], "columns": None},
            ]
        }

    def create_vendor_row(self, vendor_name: str, vendor_id: str,
                          current: float, days_1_30: float, days_31_60: float,
                          days_61_90: float, days_91_over: float, total: float) -> Dict[str, Any]:
        """Create a vendor row with aging data"""
        return {
            "id": None, "parentId": None, "header": None, "rows": None, "summary": None,
            "colData": [
                self.make_coldata_cell(vendor_name, vendor_id),
                self.make_coldata_cell(str(current) if current else ""),
                self.make_coldata_cell(str(days_1_30) if days_1_30 else ""),
                self.make_coldata_cell(str(days_31_60) if days_31_60 else ""),
                self.make_coldata_cell(str(days_61_90) if days_61_90 else ""),
                self.make_coldata_cell(str(days_91_over) if days_91_over else ""),
                self.make_coldata_cell(str(total)),
            ],
            "type": None, "group": None
        }

    def create_total_row(self, current: float, days_1_30: float, days_31_60: float,
                         days_61_90: float, days_91_over: float, total: float) -> Dict[str, Any]:
        """Create the grand total summary row"""
        return {
            "id": None, "parentId": None, "header": None, "rows": None,
            "summary": {
                "colData": [
                    self.make_coldata_cell("TOTAL"),
                    self.make_coldata_cell(str(current)),
                    self.make_coldata_cell(str(days_1_30)),
                    self.make_coldata_cell(str(days_31_60)),
                    self.make_coldata_cell(str(days_61_90)),
                    self.make_coldata_cell(str(days_91_over)),
                    self.make_coldata_cell(str(total)),
                ]
            },
            "colData": [],
            "type": "SECTION", "group": "GrandTotal"
        }

    def parse_csv(self, filepath: Path) -> Dict[str, Any]:
        """Parse CSV file and convert to QuickBooks JSON format"""
        vendors = []
        totals = {'current': 0.0, '1_30': 0.0, '31_60': 0.0, '61_90': 0.0, '91_over': 0.0, 'total': 0.0}
        vendor_id = 1

        with open(filepath, 'r', encoding='utf-8') as f:
            for _ in range(4):
                f.readline()

            reader = csv.DictReader(f)
            for row in reader:
                vendor_name = row.get('Vendor', '').strip()

                if not vendor_name or vendor_name.upper() == 'TOTAL':
                    if vendor_name and vendor_name.upper() == 'TOTAL':
                        totals['current'] = self.parse_amount(row.get('CURRENT', '0'))
                        totals['1_30'] = self.parse_amount(row.get('1 - 30', '0'))
                        totals['31_60'] = self.parse_amount(row.get('31 - 60', '0'))
                        totals['61_90'] = self.parse_amount(row.get('61 - 90', '0'))
                        totals['91_over'] = self.parse_amount(row.get('91 AND OVER', '0'))
                        totals['total'] = self.parse_amount(row.get('Total', '0'))
                    break

                current = self.parse_amount(row.get('CURRENT', '0'))
                days_1_30 = self.parse_amount(row.get('1 - 30', '0'))
                days_31_60 = self.parse_amount(row.get('31 - 60', '0'))
                days_61_90 = self.parse_amount(row.get('61 - 90', '0'))
                days_91_over = self.parse_amount(row.get('91 AND OVER', '0'))
                total = self.parse_amount(row.get('Total', '0'))

                vendor_row = self.create_vendor_row(
                    vendor_name, str(vendor_id),
                    current, days_1_30, days_31_60, days_61_90, days_91_over, total
                )
                vendors.append(vendor_row)
                vendor_id += 1

        total_row = self.create_total_row(
            totals['current'], totals['1_30'], totals['31_60'],
            totals['61_90'], totals['91_over'], totals['total']
        )
        vendors.append(total_row)

        return {
            "header": self.create_header(),
            "columns": self.create_columns(),
            "rows": {"row": vendors}
        }

    def parse_xlsx(self, filepath: Path) -> Dict[str, Any]:
        """Parse XLSX file and convert to QuickBooks JSON format"""
        self.check_xlsx_support()

        vendors = []
        totals = {'current': 0.0, '1_30': 0.0, '31_60': 0.0, '61_90': 0.0, '91_over': 0.0, 'total': 0.0}
        vendor_id = 1

        workbook = openpyxl.load_workbook(filepath, data_only=True)
        sheet = workbook.active

        header_row = None
        for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            # Look for aging bucket headers: Current, 1-30, 31-60, etc.
            row_str = ' '.join(str(cell) for cell in row if cell)
            if ('Current' in row_str and 'Total' in row_str) or \
               ('1' in row_str and '30' in row_str and '60' in row_str):
                header_row = idx
                break

        if not header_row:
            raise ValueError("Could not find header row in XLSX file")

        headers = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        col_map = {str(header).strip(): idx for idx, header in enumerate(headers) if header}

        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[col_map.get('Vendor', 0)]:
                continue

            vendor_name = str(row[col_map.get('Vendor', 0)]).strip()

            if vendor_name.upper() == 'TOTAL':
                totals['current'] = self.parse_amount(str(row[col_map.get('CURRENT', 1)] or '0'))
                totals['1_30'] = self.parse_amount(str(row[col_map.get('1 - 30', 2)] or '0'))
                totals['31_60'] = self.parse_amount(str(row[col_map.get('31 - 60', 3)] or '0'))
                totals['61_90'] = self.parse_amount(str(row[col_map.get('61 - 90', 4)] or '0'))
                totals['91_over'] = self.parse_amount(str(row[col_map.get('91 AND OVER', 5)] or '0'))
                totals['total'] = self.parse_amount(str(row[col_map.get('Total', 6)] or '0'))
                break

            current = self.parse_amount(str(row[col_map.get('CURRENT', 1)] or '0'))
            days_1_30 = self.parse_amount(str(row[col_map.get('1 - 30', 2)] or '0'))
            days_31_60 = self.parse_amount(str(row[col_map.get('31 - 60', 3)] or '0'))
            days_61_90 = self.parse_amount(str(row[col_map.get('61 - 90', 4)] or '0'))
            days_91_over = self.parse_amount(str(row[col_map.get('91 AND OVER', 5)] or '0'))
            total = self.parse_amount(str(row[col_map.get('Total', 6)] or '0'))

            vendor_row = self.create_vendor_row(
                vendor_name, str(vendor_id),
                current, days_1_30, days_31_60, days_61_90, days_91_over, total
            )
            vendors.append(vendor_row)
            vendor_id += 1

        total_row = self.create_total_row(
            totals['current'], totals['1_30'], totals['31_60'],
            totals['61_90'], totals['91_over'], totals['total']
        )
        vendors.append(total_row)

        return {
            "header": self.create_header(),
            "columns": self.create_columns(),
            "rows": {"row": vendors}
        }

    # Aging-bucket order matching the report columns (Current ... 91+ then TOTAL).
    _BUCKET_KEYS = ['current', '1_30', '31_60', '61_90', '91_over', 'total']

    # An amount: optional leading "-" / "$", thousands-separated, two decimals.
    _AMOUNT_RE = re.compile(r'^-?\$?[\d,]+\.\d{2}\$?$')

    @classmethod
    def _is_amount(cls, text: str) -> bool:
        return bool(cls._AMOUNT_RE.match(text.strip()))

    @staticmethod
    def _group_words_into_lines(words: List[Dict[str, Any]], tol: float = 3.0) -> List[List[Dict[str, Any]]]:
        """Cluster extracted words into visual lines by their vertical position."""
        lines: List[List[Dict[str, Any]]] = []
        for w in sorted(words, key=lambda x: (round(x['top']), x['x0'])):
            placed = False
            for line in lines:
                if abs(line[0]['top'] - w['top']) <= tol:
                    line.append(w)
                    placed = True
                    break
            if not placed:
                lines.append([w])
        for line in lines:
            line.sort(key=lambda x: x['x0'])
        lines.sort(key=lambda ln: ln[0]['top'])
        return lines

    def _find_bucket_edges(self, lines: List[List[Dict[str, Any]]]) -> Optional[List[float]]:
        """Locate the header row and return the right-edge x of each aging column.

        QB right-aligns every amount under its column header, so we match each
        amount to a bucket by comparing its right edge (x1) against the right
        edge of the header label for that column.
        """
        for line in lines:
            joined = ' '.join(w['text'] for w in line).upper()
            if 'VENDOR' not in joined or ('CURRENT' not in joined and 'TOTAL' not in joined):
                continue
            # Header labels wrap across multiple words (e.g. "1 - 30",
            # "91 AND OVER"). Walk left-to-right and treat the right edge of the
            # word that ends each label as that column's anchor.
            edges: List[float] = []
            labels = ['CURRENT', '1 - 30', '31 - 60', '61 - 90', '91 AND OVER', 'TOTAL']
            label_idx = 0
            buf = ''
            buf_x1 = 0.0
            for w in line:
                txt = w['text'].upper()
                if txt == 'VENDOR':
                    continue
                buf = (buf + ' ' + txt).strip() if buf else txt
                buf_x1 = w['x1']
                if label_idx < len(labels) and buf.replace(' ', '') == labels[label_idx].replace(' ', ''):
                    edges.append(buf_x1)
                    label_idx += 1
                    buf = ''
            if len(edges) == len(labels):
                return edges
        return None

    def _assign_buckets(self, amount_words: List[Dict[str, Any]], edges: List[float]) -> Dict[str, float]:
        """Map each amount to its aging bucket by nearest header right-edge."""
        result = {k: 0.0 for k in self._BUCKET_KEYS}
        for w in amount_words:
            x1 = w['x1']
            best = min(range(len(edges)), key=lambda i: abs(edges[i] - x1))
            result[self._BUCKET_KEYS[best]] = self.parse_amount(w['text'])
        return result

    def parse_pdf(self, filepath: Path) -> Dict[str, Any]:
        """Parse PDF file and convert to QuickBooks JSON format.

        QB exports border-less PDFs, so plain text extraction collapses the
        aging columns unpredictably. We instead work from word positions: amounts
        are right-aligned under their column headers, so each amount is assigned
        to a bucket by matching its right edge to the header column edges. Vendor
        names (which may wrap across words) are everything that is not an amount.
        """
        self.check_pdf_support()

        vendors: List[Dict[str, Any]] = []
        totals = {k: 0.0 for k in self._BUCKET_KEYS}
        vendor_id = 1

        print("[AP-PARSER] Starting PDF parse")

        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                words = page.extract_words()
                if not words:
                    continue

                lines = self._group_words_into_lines(words)
                edges = self._find_bucket_edges(lines)
                if not edges:
                    print("[AP-PARSER] No header found on page; skipping")
                    continue

                header_top = None
                for line in lines:
                    joined = ' '.join(w['text'] for w in line).upper()
                    if 'VENDOR' in joined and ('CURRENT' in joined or 'TOTAL' in joined):
                        header_top = line[0]['top']
                        break

                done = False
                for line in lines:
                    if header_top is not None and line[0]['top'] <= header_top:
                        continue

                    amount_words = [w for w in line if self._is_amount(w['text'])]
                    name_words = [w for w in line if not self._is_amount(w['text'])]
                    name = ' '.join(w['text'] for w in name_words).strip()

                    # Report furniture: footers ("... GMTZ"), page numbers, etc.
                    if not amount_words and not name:
                        continue
                    if 'GMTZ' in name.upper() or re.search(r'\bGMT', name.upper()):
                        continue

                    if name.upper().startswith('TOTAL') and amount_words:
                        buckets = self._assign_buckets(amount_words, edges)
                        totals.update(buckets)
                        print(f"[AP-PARSER] Found TOTAL row: {totals}")
                        done = True
                        break

                    if not amount_words:
                        # A name-only line with no amounts: skip stray furniture.
                        continue

                    buckets = self._assign_buckets(amount_words, edges)
                    vendor_row = self.create_vendor_row(
                        name, str(vendor_id),
                        buckets['current'], buckets['1_30'], buckets['31_60'],
                        buckets['61_90'], buckets['91_over'], buckets['total']
                    )
                    vendors.append(vendor_row)
                    vendor_id += 1
                    print(f"[AP-PARSER] Added vendor: {name} (${buckets['total']})")

                if done:
                    break

        print(f"[AP-PARSER] Parsed {len(vendors)} vendor rows")

        total_row = self.create_total_row(
            totals['current'], totals['1_30'], totals['31_60'],
            totals['61_90'], totals['91_over'], totals['total']
        )
        vendors.append(total_row)

        return {
            "header": self.create_header(),
            "columns": self.create_columns(),
            "rows": {"row": vendors}
        }

    def convert_to_json(self, filepath: Path, output_path: Optional[Path] = None) -> str:
        """Convert a file to JSON format"""
        import json
        result = self.convert_file(filepath)

        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, indent=2)
            vendor_count = len(result['rows']['row']) - 1
            return f"Converted {vendor_count} vendors to {output_path}"
        else:
            return json.dumps(result, indent=2)


def main():
    parser = argparse.ArgumentParser(description='Convert A/P Aging Summary reports to JSON format')
    parser.add_argument('input', help='Input file (CSV, XLSX, or PDF)')
    parser.add_argument('-o', '--output', help='Output JSON file (default: print to stdout)')
    parser.add_argument('--batch', action='store_true', help='Process all files in a directory')

    args = parser.parse_args()

    converter = AccountsPayableConverter()

    if args.batch:
        input_path = Path(args.input)
        if not input_path.is_dir():
            print(f"Error: {input_path} is not a directory", file=sys.stderr)
            sys.exit(1)

        output_dir = Path(args.output) if args.output else input_path.parent / 'converted'
        output_dir.mkdir(exist_ok=True)

        for file in input_path.glob('*'):
            if file.suffix.lower() in ['.csv', '.xlsx', '.pdf']:
                try:
                    output_file = output_dir / f"{file.stem}_ap_aging.json"
                    result = converter.convert_to_json(file, output_file)
                    print(result)
                except Exception as e:
                    print(f"Error processing {file}: {e}", file=sys.stderr)
    else:
        input_path = Path(args.input)
        if not input_path.exists():
            print(f"Error: {input_path} does not exist", file=sys.stderr)
            sys.exit(1)

        try:
            if args.output:
                result = converter.convert_to_json(input_path, Path(args.output))
                print(result)
            else:
                print(converter.convert_to_json(input_path))
        except Exception as e:
            print(f"Error: {e}", file=sys.stderr)
            sys.exit(1)


if __name__ == '__main__':
    main()
