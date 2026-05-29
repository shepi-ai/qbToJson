#!/usr/bin/env python3
"""
Trial Balance Document Converter
Converts CSV, XLSX, and PDF trial balance reports to QuickBooks JSON format
Handles monthly trial balance data with debit and credit columns
"""

import json
import csv
import sys
import os
from datetime import datetime, timezone, date
from pathlib import Path
import argparse
import re
from typing import Dict, List, Any, Optional, Tuple
import calendar

from base_converter import BaseConverter, XLSX_SUPPORT, PDF_SUPPORT

if XLSX_SUPPORT:
    import openpyxl

if PDF_SUPPORT:
    import pdfplumber


class TrialBalanceConverter(BaseConverter):
    """Converts Trial Balance documents to QuickBooks-style JSON format"""

    def __init__(self, use_account_lookup: bool = True, api_base_url: str = "http://localhost:8080"):
        super().__init__(use_account_lookup=use_account_lookup, api_base_url=api_base_url)

    def extract_date_from_as_of(self, text: str) -> Tuple[str, str, date, date]:
        """Extract date from 'As of [Date]' format"""
        # Match patterns like "As of May 31, 2024" or "As of May 30, 2025"
        match = re.search(r'As of\s+([A-Za-z]+)\s+(\d{1,2}),?\s+(\d{4})', text, re.IGNORECASE)
        if match:
            month_name = match.group(1)
            year = match.group(3)

            # Convert month name to number
            try:
                month_num = datetime.strptime(month_name, '%B').month
            except ValueError:
                try:
                    month_num = datetime.strptime(month_name[:3], '%b').month
                except ValueError:
                    month_num = 1

            # Create start and end dates for the month
            start_date = date(int(year), month_num, 1)
            last_day = calendar.monthrange(int(year), month_num)[1]
            end_date = date(int(year), month_num, last_day)

            return month_name.upper()[:3], year, start_date, end_date

        # Fallback
        return "JAN", "2025", date(2025, 1, 1), date(2025, 1, 31)

    def create_row_object(self, account_name: str, debit: str = "", credit: str = "",
                         account_id: Optional[str] = None, is_total: bool = False) -> Dict[str, Any]:
        """Create a row object for the trial balance"""
        if is_total:
            # Total row
            return {
                "id": None,
                "parentId": None,
                "header": None,
                "rows": None,
                "summary": {
                    "colData": [
                        {"attributes": None, "value": "TOTAL", "id": None, "href": None},
                        {"attributes": None, "value": debit, "id": None, "href": None},
                        {"attributes": None, "value": credit, "id": None, "href": None}
                    ]
                },
                "colData": [],
                "type": "SECTION",
                "group": "GrandTotal"
            }
        else:
            # Regular account row
            return {
                "id": None,
                "parentId": None,
                "header": None,
                "rows": None,
                "summary": None,
                "colData": [
                    {"attributes": None, "value": account_name, "id": account_id, "href": None},
                    {"attributes": None, "value": debit, "id": None, "href": None},
                    {"attributes": None, "value": credit, "id": None, "href": None}
                ],
                "type": None,
                "group": None
            }

    def parse_csv_data(self, filepath: Path) -> Dict[str, Dict[str, Any]]:
        """Parse CSV file and extract trial balance data by month"""
        data_by_month = {}

        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)

            # Find header row with months. Match cells that look like a month-year
            # column header (e.g. "January 2023"); this deliberately skips the
            # "As of December 31, 2025" title line, where the month is followed by a
            # day rather than a 4-digit year, which would otherwise be picked as the
            # header and yield zero month columns.
            month_year_re = re.compile(
                r'(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)[A-Z]*\s+\d{4}'
            )
            header_row_idx = -1
            for i, row in enumerate(rows):
                if len(row) > 1 and any(
                    cell and month_year_re.search(str(cell).upper()) for cell in row
                ):
                    header_row_idx = i
                    break

            if header_row_idx == -1:
                raise ValueError("Could not find header row with months")

            # Parse months from header
            header_row = rows[header_row_idx]
            month_columns = []

            # Find month columns (they usually come in pairs - debit and credit)
            col_idx = 1  # Skip first column (account names)
            while col_idx < len(header_row):
                cell = header_row[col_idx]
                if cell and month_year_re.search(cell.upper()):
                    month, year, start_date, end_date = self.parse_month_year(cell)
                    month_columns.append({
                        'month': month,
                        'year': year,
                        'start_date': start_date,
                        'end_date': end_date,
                        'debit_col': col_idx,
                        'credit_col': col_idx + 1 if col_idx + 1 < len(header_row) else col_idx
                    })
                    col_idx += 2  # Skip to next month (assuming debit/credit pairs)
                else:
                    col_idx += 1

            # Initialize data structure for each month
            for month_info in month_columns:
                month_key = f"{month_info['year']}-{month_info['month']}"
                data_by_month[month_key] = {
                    'month': month_info['month'],
                    'year': month_info['year'],
                    'start_date': month_info['start_date'],
                    'end_date': month_info['end_date'],
                    'accounts': [],
                    'total_debit': 0.0,
                    'total_credit': 0.0
                }

            # Parse account data
            for row_idx in range(header_row_idx + 1, len(rows)):
                row = rows[row_idx]
                if not row or not row[0] or row[0].strip().upper() in ['TOTAL', 'TOTALS', 'GRAND TOTAL']:
                    continue

                account_name = row[0].strip()
                if not account_name:
                    continue

                # Skip metadata lines
                if any(skip in account_name.lower() for skip in ['accrual basis', 'gmt', 'gmtz']):
                    continue

                # Get account ID
                account_id = self.get_or_create_account_id(account_name)

                # Extract values for each month
                for month_info in month_columns:
                    month_key = f"{month_info['year']}-{month_info['month']}"

                    # Get debit value - track if cell has explicit data
                    debit_value = 0.0
                    debit_present = False
                    if month_info['debit_col'] < len(row):
                        debit_str = row[month_info['debit_col']].strip().replace(',', '').replace('$', '').replace('(', '-').replace(')', '')
                        if debit_str and debit_str != '-':
                            debit_present = True
                            try:
                                debit_value = float(debit_str)
                            except ValueError:
                                debit_value = 0.0

                    # Get credit value - track if cell has explicit data
                    credit_value = 0.0
                    credit_present = False
                    if month_info['credit_col'] < len(row):
                        credit_str = row[month_info['credit_col']].strip().replace(',', '').replace('$', '').replace('(', '-').replace(')', '')
                        if credit_str and credit_str != '-':
                            credit_present = True
                            try:
                                credit_value = float(credit_str)
                            except ValueError:
                                credit_value = 0.0

                    # Include if any value is present (even 0) or is a special account
                    if debit_present or credit_present or account_name in ['Retained Earnings']:
                        data_by_month[month_key]['accounts'].append({
                            'name': account_name,
                            'id': account_id,
                            'debit': debit_value,
                            'credit': credit_value
                        })
                        data_by_month[month_key]['total_debit'] += debit_value
                        data_by_month[month_key]['total_credit'] += credit_value

        return data_by_month

    def parse_single_month_xlsx(self, filepath: Path) -> Dict[str, Dict[str, Any]]:
        """Parse single-month XLSX format (e.g., 'As of December 31, 2025')"""
        self.check_xlsx_support()

        # Load without data_only to get formulas, then extract values
        workbook = openpyxl.load_workbook(filepath, data_only=False)
        sheet = workbook.active

        data_by_month = {}

        # Convert to list of lists and extract values from formulas
        rows = []
        for row in sheet.iter_rows(values_only=True):
            row_data = []
            for cell in row:
                if cell is None or cell == '':
                    row_data.append('')
                else:
                    cell_str = str(cell)
                    # Check if it's an Excel formula (e.g., '=1201.00')
                    if cell_str.startswith('='):
                        # Extract number from formula
                        match = re.search(r'=([\d.]+)', cell_str)
                        if match:
                            try:
                                row_data.append(float(match.group(1)))
                            except ValueError:
                                row_data.append(cell)
                        else:
                            row_data.append(cell)
                    else:
                        row_data.append(cell)
            rows.append(row_data)

        if len(rows) < 5:
            print(f"[DEBUG] XLSX has too few rows: {len(rows)}", file=sys.stderr)
            return data_by_month

        # Find "As of [Date]" in early rows (typically row 2 or 3)
        date_text = ""
        for i in range(min(5, len(rows))):
            row_text = ' '.join(str(cell) for cell in rows[i] if cell)
            if "As of" in row_text:
                date_text = row_text
                break

        if not date_text:
            print(f"[DEBUG] Could not find 'As of' date in XLSX", file=sys.stderr)
            return data_by_month

        # Extract date from header
        month, year, start_date, end_date = self.extract_date_from_as_of(date_text)
        month_key = f"{year}-{month}"

        # Initialize month data
        data_by_month[month_key] = {
            'month': month,
            'year': year,
            'start_date': start_date,
            'end_date': end_date,
            'accounts': [],
            'total_debit': 0.0,
            'total_credit': 0.0
        }

        # Find DEBIT/CREDIT header row
        header_idx = -1
        debit_col = -1
        credit_col = -1

        for i in range(min(10, len(rows))):
            for j, cell in enumerate(rows[i]):
                cell_str = str(cell).upper()
                if 'DEBIT' in cell_str:
                    header_idx = i
                    debit_col = j
                if 'CREDIT' in cell_str:
                    credit_col = j
            if header_idx != -1 and debit_col != -1 and credit_col != -1:
                break

        if header_idx == -1:
            print(f"[DEBUG] Could not find DEBIT/CREDIT headers in XLSX", file=sys.stderr)
            return data_by_month

        print(f"[DEBUG] Found headers at row {header_idx}, debit col {debit_col}, credit col {credit_col}", file=sys.stderr)

        # Parse account data (after header, before TOTAL)
        for row_idx in range(header_idx + 1, len(rows)):
            row = rows[row_idx]

            if not row or len(row) < 2:
                continue

            # First column should be account name
            account_name = str(row[0]).strip()

            # Stop at TOTAL line
            if account_name.upper() in ['TOTAL', 'TOTALS']:
                break

            # Skip empty account names
            if not account_name or account_name == '' or account_name == 'None':
                continue

            # Skip footer lines (date stamps, etc.)
            if any(skip in account_name.lower() for skip in ['accrual basis', 'gmt', 'pm', 'am']):
                continue

            # Get debit value
            debit_value = 0.0
            if debit_col < len(row):
                debit_cell = row[debit_col]
                if debit_cell and debit_cell != '':
                    try:
                        debit_value = float(str(debit_cell).replace(',', ''))
                    except (ValueError, AttributeError):
                        debit_value = 0.0

            # Get credit value
            credit_value = 0.0
            if credit_col < len(row):
                credit_cell = row[credit_col]
                if credit_cell and credit_cell != '':
                    try:
                        credit_value = float(str(credit_cell).replace(',', ''))
                    except (ValueError, AttributeError):
                        credit_value = 0.0

            # Get account ID
            account_id = self.get_or_create_account_id(account_name)

            # Include all accounts (even zero-value) to match QB API output
            data_by_month[month_key]['accounts'].append({
                'name': account_name,
                'id': account_id,
                'debit': debit_value,
                'credit': credit_value
            })
            data_by_month[month_key]['total_debit'] += debit_value
            data_by_month[month_key]['total_credit'] += credit_value

        return data_by_month

    def parse_xlsx_data(self, filepath: Path) -> Dict[str, Dict[str, Any]]:
        """Parse XLSX file and extract trial balance data by month"""
        self.check_xlsx_support()

        # First, detect format by reading with data_only=True to evaluate formulas
        workbook = openpyxl.load_workbook(filepath, data_only=True)
        sheet = workbook.active

        # Convert first few rows to check format
        rows = []
        row_count = 0
        for row in sheet.iter_rows(values_only=True):
            rows.append([cell if cell is not None else '' for cell in row])
            row_count += 1
            if row_count >= 10:  # Only need first 10 rows for detection
                break

        # Check if it's single-month format
        has_as_of = False
        for i in range(min(5, len(rows))):
            row_text = ' '.join(str(cell) for cell in rows[i] if cell)
            if "As of" in row_text:
                has_as_of = True
                break

        # Count month columns with years in remaining rows
        # Look at all cells in header rows (typically row 4-5) for month-year patterns
        month_year_count = 0
        for i in range(4, min(len(rows), 7)):  # Check rows 4-6 (header area)
            for cell in rows[i]:
                cell_text = str(cell).upper()
                # Match patterns like "JAN 2022", "JANUARY 2022", "Jan 2025"
                if re.search(r'(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)[A-Z]*\s+\d{4}', cell_text):
                    month_year_count += 1

        print(f"[DEBUG] Format detection: has_as_of={has_as_of}, month_year_count={month_year_count}", file=sys.stderr)

        if has_as_of and month_year_count < 2:
            print(f"[DEBUG] Detected single-month XLSX format (has 'As of' and fewer than 2 month columns)", file=sys.stderr)
            return self.parse_single_month_xlsx(filepath)

        # Fall back to multi-month parser
        print(f"[DEBUG] Using multi-month XLSX parser", file=sys.stderr)

        # Re-load workbook without data_only for multi-month parsing
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active

        # Convert to list of lists, preserving None vs 0 distinction
        rows = []
        raw_rows = []  # Keep original values to distinguish None from 0
        for row in sheet.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else '' for cell in row])
            raw_rows.append(list(row))

        # Find header row with months
        header_row_idx = -1
        for i, row in enumerate(rows):
            if len(row) > 1:
                # Skip rows that are likely report headers (e.g., "As of July 31, 2025")
                if i < 3 and any(phrase in ' '.join(str(cell) for cell in row if cell).lower() for phrase in ['as of', 'trial balance', 'company']):
                    continue

                row_text = ' '.join(str(cell) for cell in row if cell)
                # Look for rows with multiple month names (indicating column headers)
                month_count = 0
                for month in ['JANUARY', 'FEBRUARY', 'MARCH', 'APRIL', 'MAY', 'JUNE', 'JULY', 'AUGUST', 'SEPTEMBER', 'OCTOBER', 'NOVEMBER', 'DECEMBER', 'JAN', 'FEB', 'MAR', 'APR', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']:
                    if month in row_text.upper():
                        # Check if it's followed by a year (to confirm it's a month header)
                        if re.search(rf'{month}\s+\d{{4}}', row_text.upper()):
                            month_count += 1

                # If we found at least 2 months with years, this is likely our header row
                if month_count >= 2:
                    header_row_idx = i
                    break

        if header_row_idx == -1:
            raise ValueError("Could not find header row with months")

        # Parse months from header
        header_row = rows[header_row_idx]
        month_columns = []

        # Find month columns
        col_idx = 1
        while col_idx < len(header_row):
            cell = header_row[col_idx]
            if cell and any(month in str(cell).upper() for month in ['JANUARY', 'FEBRUARY', 'MARCH', 'APRIL', 'MAY', 'JUNE', 'JULY', 'AUGUST', 'SEPTEMBER', 'OCTOBER', 'NOVEMBER', 'DECEMBER', 'JAN', 'FEB', 'MAR', 'APR', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']):
                month, year, start_date, end_date = self.parse_month_year(str(cell))

                # Check if next columns are labeled as Debit/Credit
                has_labels = False
                if header_row_idx + 1 < len(rows):
                    next_row = rows[header_row_idx + 1]
                    if col_idx < len(next_row) and 'DEBIT' in next_row[col_idx].upper():
                        has_labels = True

                month_columns.append({
                    'month': month,
                    'year': year,
                    'start_date': start_date,
                    'end_date': end_date,
                    'debit_col': col_idx,
                    'credit_col': col_idx + 1 if col_idx + 1 < len(header_row) else col_idx,
                    'has_labels': has_labels
                })
                col_idx += 2
            else:
                col_idx += 1

        # Initialize data structure
        data_by_month = {}
        for month_info in month_columns:
            month_key = f"{month_info['year']}-{month_info['month']}"
            data_by_month[month_key] = {
                'month': month_info['month'],
                'year': month_info['year'],
                'start_date': month_info['start_date'],
                'end_date': month_info['end_date'],
                'accounts': [],
                'total_debit': 0.0,
                'total_credit': 0.0
            }

        # Skip label row if present
        data_start_row = header_row_idx + 2 if any(m['has_labels'] for m in month_columns) else header_row_idx + 1

        # Parse account data
        for row_idx in range(data_start_row, len(rows)):
            row = rows[row_idx]
            raw_row = raw_rows[row_idx] if row_idx < len(raw_rows) else []
            if not row or not row[0] or row[0].strip().upper() in ['TOTAL', 'TOTALS', 'GRAND TOTAL']:
                continue

            account_name = row[0].strip()
            if not account_name:
                continue

            # Skip metadata lines (e.g., "Accrual Basis Wednesday, March 04, 2026...")
            if any(skip in account_name.lower() for skip in ['accrual basis', 'gmt', 'gmtz']):
                continue

            # Get account ID
            account_id = self.get_or_create_account_id(account_name)

            # Extract values for each month
            for month_info in month_columns:
                month_key = f"{month_info['year']}-{month_info['month']}"

                # Check raw cell values to distinguish None (absent) from 0 (explicit zero)
                raw_debit = raw_row[month_info['debit_col']] if month_info['debit_col'] < len(raw_row) else None
                raw_credit = raw_row[month_info['credit_col']] if month_info['credit_col'] < len(raw_row) else None
                debit_present = raw_debit is not None
                credit_present = raw_credit is not None

                # Get debit value
                debit_value = 0.0
                if debit_present and month_info['debit_col'] < len(row):
                    debit_str = row[month_info['debit_col']].strip().replace(',', '').replace('$', '').replace('(', '-').replace(')', '')
                    if debit_str.startswith('='):
                        if '=' in debit_str and any(c.isdigit() for c in debit_str):
                            numbers = re.findall(r'[\d.]+', debit_str)
                            if numbers:
                                try:
                                    debit_value = float(numbers[0])
                                except ValueError:
                                    debit_value = 0.0
                    else:
                        try:
                            debit_value = float(debit_str) if debit_str and debit_str != '-' else 0.0
                        except ValueError:
                            debit_value = 0.0

                # Get credit value
                credit_value = 0.0
                if credit_present and month_info['credit_col'] < len(row):
                    credit_str = row[month_info['credit_col']].strip().replace(',', '').replace('$', '').replace('(', '-').replace(')', '')
                    if credit_str.startswith('='):
                        if '=' in credit_str and any(c.isdigit() for c in credit_str):
                            numbers = re.findall(r'[\d.]+', credit_str)
                            if numbers:
                                try:
                                    credit_value = float(numbers[0])
                                except ValueError:
                                    credit_value = 0.0
                    else:
                        try:
                            credit_value = float(credit_str) if credit_str and credit_str != '-' else 0.0
                        except ValueError:
                            credit_value = 0.0

                # Only include if at least one cell has data (not None in raw XLSX)
                if debit_present or credit_present:
                    data_by_month[month_key]['accounts'].append({
                        'name': account_name,
                        'id': account_id,
                        'debit': debit_value,
                        'credit': credit_value
                    })
                    data_by_month[month_key]['total_debit'] += debit_value
                    data_by_month[month_key]['total_credit'] += credit_value

        return data_by_month

    def parse_single_month_pdf(self, filepath: Path) -> Dict[str, Dict[str, Any]]:
        """Parse single-month PDF format (e.g., 'As of May 31, 2024')"""
        self.check_pdf_support()

        data_by_month = {}

        with pdfplumber.open(filepath) as pdf:
            # Extract text from all pages
            full_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"

            if not full_text:
                return data_by_month

            # Extract date from "As of [Date]" format
            month, year, start_date, end_date = self.extract_date_from_as_of(full_text)
            month_key = f"{year}-{month}"

            # Initialize month data
            data_by_month[month_key] = {
                'month': month,
                'year': year,
                'start_date': start_date,
                'end_date': end_date,
                'accounts': [],
                'total_debit': 0.0,
                'total_credit': 0.0
            }

            lines = full_text.split('\n')

            # Find DEBIT/CREDIT header line
            header_idx = -1
            for i, line in enumerate(lines):
                if 'DEBIT' in line.upper() and 'CREDIT' in line.upper():
                    header_idx = i
                    break

            if header_idx == -1:
                print(f"[DEBUG] Could not find DEBIT/CREDIT header in PDF", file=sys.stderr)
                return data_by_month

            # Parse account lines (after header, before TOTAL)
            for line_idx in range(header_idx + 1, len(lines)):
                line = lines[line_idx].strip()

                # Stop at TOTAL line
                if not line or line.upper().startswith('TOTAL'):
                    break

                # Skip page headers that repeat
                if any(skip in line.upper() for skip in ['TRIAL BALANCE', 'AS OF', 'ACCRUAL BASIS', 'DEBIT', 'CREDIT']):
                    continue

                # Extract account name and values using regex
                # Pattern: account name followed by numbers
                # Account names can contain letters, spaces, parentheses, slashes, colons, etc.
                match = re.match(r'^(.+?)\s+([\d,]+\.?\d*)\s*([\d,]+\.?\d*)?$', line)

                if match:
                    account_name = match.group(1).strip()

                    # Skip if it looks like a page number or date
                    if account_name.isdigit() or re.match(r'^\d+/\d+$', account_name):
                        continue

                    # Parse debit value
                    debit_str = match.group(2).strip().replace(',', '')
                    try:
                        debit_value = float(debit_str)
                    except ValueError:
                        debit_value = 0.0

                    # Parse credit value (might be empty)
                    credit_value = 0.0
                    if match.group(3):
                        credit_str = match.group(3).strip().replace(',', '')
                        try:
                            credit_value = float(credit_str)
                        except ValueError:
                            credit_value = 0.0

                    # Determine if value is debit or credit based on account type
                    # If only one value present, infer from account name
                    if debit_value > 0 and credit_value == 0:
                        # Check if this is likely a credit account
                        if any(keyword in account_name.upper() for keyword in
                               ['PAYABLE', 'EQUITY', 'EARNINGS', 'LOAN', 'RETAINED', 'CONTRIBUTIONS', 'REVENUE', 'INCOME', 'SALES', 'SERVICES']):
                            credit_value = debit_value
                            debit_value = 0.0

                    # Get account ID
                    account_id = self.get_or_create_account_id(account_name)

                    # Include all accounts (even zero-value) to match QB API output
                    data_by_month[month_key]['accounts'].append({
                        'name': account_name,
                        'id': account_id,
                        'debit': debit_value,
                        'credit': credit_value
                    })
                    data_by_month[month_key]['total_debit'] += debit_value
                    data_by_month[month_key]['total_credit'] += credit_value

        return data_by_month

    # ------------------------------------------------------------------
    # PDF format detection
    # ------------------------------------------------------------------
    @staticmethod
    def _count_pdf_month_columns(first_page_text: str) -> int:
        """Count distinct month-year column headers in the PDF header line.

        QB wide trial balances use full month names ("JANUARY 2023") in the
        header row, while single-month reports only carry an "As of <date>"
        line (no month-year column headers). This counts month-year tokens so
        detection can route multi-month PDFs to the columnar parser.
        """
        month_year_pattern = (
            r'(?:JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)[A-Z]*\s+\d{4}'
        )
        return len(re.findall(month_year_pattern, first_page_text.upper()))

    def parse_pdf_data(self, filepath: Path) -> Dict[str, Dict[str, Any]]:
        """Parse PDF file and extract trial balance data by month"""
        self.check_pdf_support()

        # First, detect format
        with pdfplumber.open(filepath) as pdf:
            first_page_text = pdf.pages[0].extract_text() if pdf.pages else ""

        # Single-month reports carry "As of <date>" plus DEBIT/CREDIT but have
        # fewer than 2 month-year column headers. Multi-month wide reports have
        # many month-year headers (e.g. "JANUARY 2023 FEBRUARY 2023 ...").
        month_col_count = self._count_pdf_month_columns(first_page_text)
        has_debit_credit = (
            "DEBIT" in first_page_text.upper() and "CREDIT" in first_page_text.upper()
        )

        if month_col_count < 2 and "As of" in first_page_text and has_debit_credit:
            print(f"[DEBUG] Detected single-month PDF format", file=sys.stderr)
            return self.parse_single_month_pdf(filepath)

        # Multi-month columnar parser (word-position / x-snapping based).
        print(
            f"[DEBUG] Using multi-month PDF parser (month_col_count={month_col_count})",
            file=sys.stderr,
        )
        return self._parse_multi_month_pdf(filepath)

    # ------------------------------------------------------------------
    # Multi-month wide trial balance PDF parser
    # ------------------------------------------------------------------
    @staticmethod
    def _cluster_words_into_rows(words: List[Dict[str, Any]], tol: float = 3.0) -> List[List[Dict[str, Any]]]:
        """Group extracted words into visual rows by their 'top' coordinate."""
        rows: List[List[Dict[str, Any]]] = []
        for w in sorted(words, key=lambda w: (w['top'], w['x0'])):
            placed = False
            for row in rows:
                if abs(row[0]['top'] - w['top']) <= tol:
                    row.append(w)
                    placed = True
                    break
            if not placed:
                rows.append([w])
        # Sort each row left-to-right and order rows top-to-bottom
        for row in rows:
            row.sort(key=lambda w: w['x0'])
        rows.sort(key=lambda r: min(w['top'] for w in r))
        return rows

    @staticmethod
    def _is_amount_token(text: str) -> bool:
        """Return True if a token is a numeric amount (handles commas/parens/$)."""
        cleaned = text.strip().replace(',', '').replace('$', '')
        cleaned = cleaned.replace('(', '').replace(')', '').replace('-', '', 1)
        if cleaned in ('', '.'):
            return False
        return bool(re.fullmatch(r'\d*\.?\d+', cleaned))

    @staticmethod
    def _parse_amount(text: str) -> float:
        """Parse a numeric amount token (parentheses => negative)."""
        t = text.strip()
        negative = t.startswith('(') and t.endswith(')')
        cleaned = t.replace(',', '').replace('$', '').replace('(', '').replace(')', '')
        try:
            val = float(cleaned)
        except ValueError:
            return 0.0
        return -val if negative else val

    def _parse_multi_month_pdf(self, filepath: Path) -> Dict[str, Dict[str, Any]]:
        """Parse a wide multi-month QB trial balance PDF.

        The report is split into horizontal page-groups (each covering a subset
        of months for all accounts) and accounts within a group can also span
        multiple pages vertically. The same accounts repeat across page-groups,
        so months are stitched together by aggregating per-month data keyed by
        (year, month). Each month has a DEBIT and a CREDIT sub-column; amounts
        are snapped to columns by their right edge (x1) versus the per-page
        DEBIT/CREDIT header positions.
        """
        data_by_month: Dict[str, Dict[str, Any]] = {}
        # Track which accounts already recorded for each month to avoid dupes
        # when an account repeats across pages/page-groups.
        seen: Dict[str, set] = {}

        # Page furniture lines to ignore. Matched as whole-line/prefix patterns
        # (NOT substrings) so legitimate account names that merely contain a
        # word like "Balance" (e.g. "Opening Balance Equity") are NOT dropped.
        def _is_furniture(name: str) -> bool:
            up = name.strip().upper()
            if up in ('TRIAL BALANCE', 'FULL NAME', 'DEBIT', 'CREDIT',
                      'DEBIT CREDIT'):
                return True
            if up.startswith('AS OF '):
                return True
            if up.startswith('ACCRUAL'):
                # Footer "Accrual Basis <date> GMTZ" line. The name-column words
                # are typically just "Accrual" (the rest sits in the number
                # columns). No real account name begins with "Accrual".
                return True
            # Footer date stamp lines and timezone markers.
            if 'GMTZ' in up:
                return True
            return False

        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
                if not words:
                    continue

                rows = self._cluster_words_into_rows(words)

                # Locate the month-header row and the DEBIT/CREDIT row.
                month_row = None
                dc_row = None
                dc_row_index = -1
                for idx, row in enumerate(rows):
                    texts_upper = [w['text'].upper() for w in row]
                    joined = ' '.join(texts_upper)
                    if month_row is None and re.search(
                        r'(JANUARY|FEBRUARY|MARCH|APRIL|MAY|JUNE|JULY|AUGUST|'
                        r'SEPTEMBER|OCTOBER|NOVEMBER|DECEMBER)\s+\d{4}', joined
                    ):
                        month_row = row
                    if dc_row is None and 'DEBIT' in texts_upper and 'CREDIT' in texts_upper:
                        dc_row = row
                        dc_row_index = idx
                    if month_row is not None and dc_row is not None:
                        break

                if month_row is None or dc_row is None:
                    continue

                # Build ordered month list from month-header row. Each month is
                # "<MonthName> <Year>" formed from adjacent tokens; record its
                # horizontal center so we can map DEBIT/CREDIT columns to it.
                months: List[Dict[str, Any]] = []
                mw = month_row
                i = 0
                while i < len(mw):
                    txt = mw[i]['text']
                    if re.fullmatch(
                        r'(?i)(JANUARY|FEBRUARY|MARCH|APRIL|MAY|JUNE|JULY|AUGUST|'
                        r'SEPTEMBER|OCTOBER|NOVEMBER|DECEMBER)', txt
                    ) and i + 1 < len(mw) and re.fullmatch(r'\d{4}', mw[i + 1]['text']):
                        month, year, start_date, end_date = self.parse_month_year(
                            f"{txt} {mw[i + 1]['text']}"
                        )
                        center = (mw[i]['x0'] + mw[i + 1]['x1']) / 2.0
                        months.append({
                            'month': month, 'year': year,
                            'start_date': start_date, 'end_date': end_date,
                            'center': center,
                        })
                        i += 2
                    else:
                        i += 1

                if not months:
                    continue

                # Build DEBIT/CREDIT column edges (use right edge x1 for snapping
                # since amounts are right-aligned).
                debit_cols = sorted(
                    [w['x1'] for w in dc_row if w['text'].upper() == 'DEBIT']
                )
                credit_cols = sorted(
                    [w['x1'] for w in dc_row if w['text'].upper() == 'CREDIT']
                )

                # Each month owns one debit edge + one credit edge, in order.
                # Map them positionally (debit/credit alternate left-to-right).
                columns = []  # list of (x1_edge, month_index, is_debit)
                for m_idx in range(len(months)):
                    if m_idx < len(debit_cols):
                        columns.append((debit_cols[m_idx], m_idx, True))
                    if m_idx < len(credit_cols):
                        columns.append((credit_cols[m_idx], m_idx, False))

                if not columns:
                    continue

                # Initialise per-month storage.
                for m in months:
                    key = f"{m['year']}-{m['month']}"
                    if key not in data_by_month:
                        data_by_month[key] = {
                            'month': m['month'], 'year': m['year'],
                            'start_date': m['start_date'], 'end_date': m['end_date'],
                            'accounts': [], 'total_debit': 0.0, 'total_credit': 0.0,
                        }
                        seen[key] = set()

                # Right edges of every numeric column, used to recognise amount
                # tokens by snapping their right edge (x1). This avoids a fixed
                # name/amount x-boundary, which mis-classifies wide 7-digit
                # values (e.g. "1,106,897.08") whose left edge creeps into the
                # name zone.
                col_edges = [c[0] for c in columns]
                snap_tol = 6.0  # right edges are very stable per page

                def _classify_row(row):
                    """Split a visual row into (amount_words, name_words).

                    A token is an amount if it parses as a number AND its right
                    edge snaps to a numeric-column right edge. Everything to the
                    left of the first amount token is the account name.
                    """
                    amts = []
                    for w in row:
                        if self._is_amount_token(w['text']):
                            nearest = min(col_edges, key=lambda e: abs(e - w['x1']))
                            if abs(nearest - w['x1']) <= snap_tol:
                                amts.append(w)
                    if amts:
                        first_amt_x0 = min(w['x0'] for w in amts)
                        names = [
                            w for w in row
                            if w['x1'] <= first_amt_x0 and w not in amts
                        ]
                    else:
                        names = list(row)
                    return amts, names

                # ---- Step 1: turn raw visual rows into clean account rows. ----
                # Each entry: {'name': str, 'amounts': [word, ...], 'top': float}.
                # QB wraps long account names onto a second visual line at the
                # same left margin with a SMALLER vertical gap than a normal
                # inter-row gap. We detect wraps by comparing each name-only
                # row's gap-to-previous against the page's typical row gap, and
                # merge the fragment into the preceding account name. A blank
                # account (name with no amounts here, but values elsewhere) has
                # a normal gap and is treated as its own row.
                data_rows = []  # list of dict(name, name_words, amount_words, top)
                for row in rows[dc_row_index + 1:]:
                    amount_words, name_words = _classify_row(row)
                    name_text = ' '.join(w['text'] for w in name_words).strip()
                    upper_name = name_text.upper()

                    if upper_name.startswith('TOTAL'):
                        break
                    if not name_text and not amount_words:
                        continue
                    if _is_furniture(name_text):
                        continue
                    if not amount_words and re.fullmatch(r'\d+/\d+', name_text):
                        continue
                    if not name_text:
                        # Amounts with no name column text -> cannot attribute.
                        continue

                    data_rows.append({
                        'name': name_text,
                        'amounts': amount_words,
                        'top': min(w['top'] for w in name_words),
                    })

                # Determine the typical (normal) inter-row gap for this page.
                gaps = [
                    data_rows[i]['top'] - data_rows[i - 1]['top']
                    for i in range(1, len(data_rows))
                ]
                gaps_sorted = sorted(g for g in gaps if g > 0)
                normal_gap = (
                    gaps_sorted[len(gaps_sorted) // 2] if gaps_sorted else 12.7
                )

                # Merge wrapped continuation lines into their parent.
                merged_rows = []
                for i, dr in enumerate(data_rows):
                    is_wrap = (
                        i > 0
                        and not dr['amounts']                       # fragment has no values
                        and (dr['top'] - data_rows[i - 1]['top']) < normal_gap * 0.92
                        and len(merged_rows) > 0
                    )
                    if is_wrap:
                        merged_rows[-1]['name'] = (
                            merged_rows[-1]['name'] + ' ' + dr['name']
                        ).strip()
                        merged_rows[-1]['amounts'].extend(dr['amounts'])
                    else:
                        merged_rows.append({
                            'name': dr['name'],
                            'amounts': list(dr['amounts']),
                        })

                # ---- Step 2: snap amounts to columns and record per month. ----
                for mr in merged_rows:
                    account_name = mr['name']
                    account_id = self.get_or_create_account_id(account_name)

                    # Snap each amount to its nearest column by right edge (x1).
                    # Track which columns actually had an amount token (a printed
                    # "0.00" counts as present, matching CSV cell semantics).
                    per_month = {m_idx: {'debit': 0.0, 'credit': 0.0,
                                         'present': False}
                                 for m_idx in range(len(months))}
                    for aw in mr['amounts']:
                        x1 = aw['x1']
                        best = min(columns, key=lambda c: abs(c[0] - x1))
                        # Reject if no column is reasonably close (>30pt off).
                        if abs(best[0] - x1) > 30:
                            continue
                        _, m_idx, is_debit = best
                        val = self._parse_amount(aw['text'])
                        per_month[m_idx]['present'] = True
                        if is_debit:
                            per_month[m_idx]['debit'] += val
                        else:
                            per_month[m_idx]['credit'] += val

                    # Record per month (dedup by account name within the month).
                    for m_idx, m in enumerate(months):
                        key = f"{m['year']}-{m['month']}"
                        if account_name in seen[key]:
                            continue
                        # Only record accounts whose cell is present for this
                        # month (matches CSV behaviour of including present cells,
                        # including explicit 0.00).
                        if not per_month[m_idx]['present']:
                            continue
                        debit_value = per_month[m_idx]['debit']
                        credit_value = per_month[m_idx]['credit']
                        data_by_month[key]['accounts'].append({
                            'name': account_name,
                            'id': account_id,
                            'debit': debit_value,
                            'credit': credit_value,
                        })
                        data_by_month[key]['total_debit'] += debit_value
                        data_by_month[key]['total_credit'] += credit_value
                        seen[key].add(account_name)

        return data_by_month

    def build_trial_balance_json(self, data_by_month: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
        """Build the complete trial balance JSON structure"""
        monthly_reports = []

        # Sort months chronologically by start_date
        sorted_months = sorted(data_by_month.keys(), key=lambda k: data_by_month[k]['start_date'])

        for month_key in sorted_months:
            month_data = data_by_month[month_key]

            # Create report structure
            report = self.create_report_structure(month_data)

            monthly_reports.append({
                "month": month_data['month'],
                "year": month_data['year'],
                "startDate": month_data['start_date'].strftime('%Y-%m-%d'),
                "endDate": month_data['end_date'].strftime('%Y-%m-%d'),
                "report": report
            })

        # Create summary
        if monthly_reports:
            summary = {
                "requestStartDate": monthly_reports[0]['startDate'],
                "requestEndDate": monthly_reports[-1]['endDate'],
                "totalMonths": len(monthly_reports),
                "accountingMethod": "Accrual"
            }
        else:
            summary = {
                "requestStartDate": "2025-01-01",
                "requestEndDate": "2025-12-31",
                "totalMonths": 0,
                "accountingMethod": "Accrual"
            }

        return {
            "monthlyReports": monthly_reports,
            "summary": summary
        }

    def create_report_structure(self, month_data: Dict[str, Any]) -> Dict[str, Any]:
        """Create the report structure for a single month"""
        timestamp = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.000+00:00')
        has_data = len(month_data['accounts']) > 0

        # Create column structure
        columns = {
            "column": [
                {
                    "colTitle": "",
                    "colType": "Account",
                    "metaData": [],
                    "columns": None
                },
                {
                    "colTitle": f"{month_data['month']} {month_data['year']}",
                    "colType": "Money",
                    "metaData": [],
                    "columns": {
                        "column": [
                            {
                                "colTitle": "Debit",
                                "colType": "Money",
                                "metaData": [
                                    {"name": "StartDate", "value": month_data['start_date'].strftime('%Y-%m-%d')},
                                    {"name": "EndDate", "value": month_data['end_date'].strftime('%Y-%m-%d')}
                                ],
                                "columns": None
                            },
                            {
                                "colTitle": "Credit",
                                "colType": "Money",
                                "metaData": [
                                    {"name": "StartDate", "value": month_data['start_date'].strftime('%Y-%m-%d')},
                                    {"name": "EndDate", "value": month_data['end_date'].strftime('%Y-%m-%d')}
                                ],
                                "columns": None
                            }
                        ]
                    }
                }
            ]
        }

        # Create rows
        rows = []

        if has_data:
            # Add account rows
            for account in month_data['accounts']:
                rows.append(self.create_row_object(
                    account['name'],
                    f"{account['debit']:.2f}" if account['debit'] != 0 else "",
                    f"{account['credit']:.2f}" if account['credit'] != 0 else "",
                    account['id']
                ))
        else:
            # No data - add Retained Earnings with empty values
            rows.append(self.create_row_object(
                "Retained Earnings",
                "",
                "",
                self.get_or_create_account_id("Retained Earnings")
            ))

        # Add total row
        rows.append(self.create_row_object(
            "",
            f"{month_data['total_debit']:.2f}",
            f"{month_data['total_credit']:.2f}",
            is_total=True
        ))

        # Create report structure
        report = {
            "header": {
                "time": timestamp,
                "reportName": "TrialBalance",
                "dateMacro": None,
                "reportBasis": "ACCRUAL",
                "startPeriod": month_data['start_date'].strftime('%Y-%m-%d'),
                "endPeriod": month_data['end_date'].strftime('%Y-%m-%d'),
                "summarizeColumnsBy": "Month",
                "currency": "USD",
                "customer": None,
                "vendor": None,
                "employee": None,
                "item": None,
                "clazz": None,
                "department": None,
                "option": [
                    {"name": "NoReportData", "value": "false" if has_data else "true"}
                ]
            },
            "columns": columns,
            "rows": {"row": rows}
        }

        return report

    # ──────────────────────────────────────────────
    # Override base class abstract methods and file dispatch
    # ──────────────────────────────────────────────

    def parse_csv(self, filepath: Path) -> Dict[str, Any]:
        """Parse CSV and return trial balance JSON structure."""
        data_by_month = self.parse_csv_data(filepath)
        return self.build_trial_balance_json(data_by_month)

    def parse_xlsx(self, filepath: Path) -> Dict[str, Any]:
        """Parse XLSX and return trial balance JSON structure."""
        data_by_month = self.parse_xlsx_data(filepath)
        return self.build_trial_balance_json(data_by_month)

    def parse_pdf(self, filepath: Path) -> Dict[str, Any]:
        """Parse PDF and return trial balance JSON structure."""
        data_by_month = self.parse_pdf_data(filepath)
        return self.build_trial_balance_json(data_by_month)

    def convert_to_json(self, filepath: Path, output_path: Optional[Path] = None) -> str:
        """Convert a file to JSON format"""
        try:
            trial_balance_data = self.convert_file(filepath)

            if output_path:
                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(trial_balance_data, f, indent=2)
                return f"Converted trial balance with {len(trial_balance_data['monthlyReports'])} monthly reports to {output_path}"
            else:
                return json.dumps(trial_balance_data, indent=2)
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise


def main():
    parser = argparse.ArgumentParser(description='Convert trial balance documents to JSON format')
    parser.add_argument('input', help='Input file (CSV, XLSX, or PDF)')
    parser.add_argument('-o', '--output', help='Output JSON file (default: print to stdout)')

    args = parser.parse_args()

    converter = TrialBalanceConverter()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: {input_path} does not exist", file=sys.stderr)
        sys.exit(1)

    try:
        if args.output:
            result = converter.convert_to_json(input_path, Path(args.output))
            print(result)
        else:
            print(converter.convert_to_json(input_path))
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
