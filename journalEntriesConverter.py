import json
import csv
import re
from datetime import datetime, timedelta
from base_converter import BaseConverter, XLSX_SUPPORT, PDF_SUPPORT

# Conditional imports handled by base_converter guards
if XLSX_SUPPORT:
    import openpyxl
    from openpyxl.utils import get_column_letter

if PDF_SUPPORT:
    import pdfplumber

from account_lookup_client import AccountLookupClient

class JournalEntriesConverter(BaseConverter):
    def __init__(self, use_account_lookup=True, api_base_url=None):
        """
        Initialize converter with optional account lookup functionality.

        :param use_account_lookup: Boolean to enable/disable account lookup
        :param api_base_url: Optional API base URL for account lookup service
        """
        super().__init__(use_account_lookup=True, api_base_url=api_base_url)
        # JE converter maintains its own account lookup via AccountLookupClient directly
        self.use_account_lookup = use_account_lookup
        if use_account_lookup:
            self.lookup_client = AccountLookupClient(api_base_url=api_base_url)
        self.account_cache = {}
        self.transactions = []

    def convert(self, file_path):
        """
        Convert journal entries file to QuickBooks JSON format.
        Backwards-compatible alias for convert_file().

        :param file_path: Path to the input file
        :return: List of journal entry dictionaries
        """
        return self.convert_file(file_path)

    def parse_csv(self, file_path):
        """Parse CSV format journal entries"""
        with open(file_path, 'r', encoding='utf-8-sig') as file:
            # Read all lines
            lines = file.readlines()

            # Skip header lines
            start_idx = 0
            for i, line in enumerate(lines):
                if 'Transaction date' in line or 'TRANSACTION DATE' in line:
                    start_idx = i
                    break

            # Process data
            current_transaction = None
            current_id = None

            for i in range(start_idx + 1, len(lines)):
                line = lines[i].strip()
                if not line or line.startswith('Total for') or line.startswith('TOTAL'):
                    if current_transaction:
                        self.transactions.append(current_transaction)
                        current_transaction = None
                        current_id = None
                    continue

                # Split line by comma, handling quoted values
                reader = csv.reader([line])
                parts = list(reader)[0]

                if len(parts) < 8:
                    continue

                # Check if this is a transaction ID line (first column has a number)
                if parts[0].strip().isdigit():
                    if current_transaction:
                        self.transactions.append(current_transaction)
                    current_id = parts[0].strip()
                    current_transaction = {
                        'id': current_id,
                        'lines': []
                    }
                    continue

                # Parse transaction line
                if current_transaction:
                    date_str = parts[1].strip()
                    trans_type = parts[2].strip()
                    num = parts[3].strip()
                    name = parts[4].strip()
                    memo = parts[5].strip()
                    account = parts[6].strip()
                    debit = parts[7].strip().replace(',', '').replace('$', '')
                    credit = parts[8].strip().replace(',', '').replace('$', '') if len(parts) > 8 else ''

                    if date_str and trans_type:
                        # Update transaction header info
                        if not current_transaction.get('date'):
                            current_transaction['date'] = date_str
                            current_transaction['type'] = trans_type
                            current_transaction['num'] = num
                            current_transaction['name'] = name
                            current_transaction['memo'] = memo

                        # Add line item
                        line_item = {
                            'account': account,
                            'description': memo,
                            'debit': float(debit) if debit else 0.0,
                            'credit': float(credit) if credit else 0.0
                        }

                        if line_item['debit'] > 0 or line_item['credit'] > 0:
                            current_transaction['lines'].append(line_item)

            # Don't forget the last transaction
            if current_transaction:
                self.transactions.append(current_transaction)

        return self.build_json_structure()

    def parse_xlsx(self, file_path):
        """Parse XLSX format journal entries"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        # Find header row
        header_row = None
        for row in range(1, min(20, sheet.max_row + 1)):
            cell_value = str(sheet.cell(row=row, column=2).value or '').upper()
            if 'TRANSACTION DATE' in cell_value:
                header_row = row
                break

        if not header_row:
            raise ValueError("Could not find header row in XLSX file")

        # Process data
        current_transaction = None
        current_id = None

        for row in range(header_row + 1, sheet.max_row + 1):
            # Get values from row
            id_val = sheet.cell(row=row, column=1).value
            date_val = sheet.cell(row=row, column=2).value
            type_val = sheet.cell(row=row, column=3).value
            num_val = sheet.cell(row=row, column=4).value
            name_val = sheet.cell(row=row, column=5).value
            memo_val = sheet.cell(row=row, column=6).value
            account_val = sheet.cell(row=row, column=7).value
            debit_val = sheet.cell(row=row, column=8).value
            credit_val = sheet.cell(row=row, column=9).value

            # Check for transaction ID
            if id_val and str(id_val).strip().isdigit():
                if current_transaction:
                    self.transactions.append(current_transaction)
                current_id = str(id_val).strip()
                current_transaction = {
                    'id': current_id,
                    'lines': []
                }
                continue

            # Check for total line
            if account_val and str(account_val).startswith('Total for'):
                if current_transaction:
                    self.transactions.append(current_transaction)
                    current_transaction = None
                    current_id = None
                continue

            # Parse transaction line
            if current_transaction and date_val:
                # Update transaction header info
                if not current_transaction.get('date'):
                    if isinstance(date_val, datetime):
                        current_transaction['date'] = date_val.strftime('%m/%d/%Y')
                    else:
                        current_transaction['date'] = str(date_val)
                    current_transaction['type'] = str(type_val or '')
                    current_transaction['num'] = str(num_val or '')
                    current_transaction['name'] = str(name_val or '')
                    current_transaction['memo'] = str(memo_val or '')

                # Add line item
                if account_val:
                    debit = float(debit_val) if debit_val else 0.0
                    credit = float(credit_val) if credit_val else 0.0

                    if debit > 0 or credit > 0:
                        line_item = {
                            'account': str(account_val),
                            'description': str(memo_val or ''),
                            'debit': debit,
                            'credit': credit
                        }
                        current_transaction['lines'].append(line_item)

        # Don't forget the last transaction
        if current_transaction:
            self.transactions.append(current_transaction)

        return self.build_json_structure()

    def parse_pdf(self, file_path):
        """Parse PDF format journal entries - pdfplumber single-space format"""
        all_text = []

        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    all_text.append(text)

        full_text = '\n'.join(all_text)
        lines = full_text.split('\n')

        print(f"[JOURNAL-PARSER] PDF has {len(lines)} lines total")

        # Find header line
        start_idx = 0
        for i, line in enumerate(lines):
            if 'TRANSACTION DATE' in line.upper():
                start_idx = i + 1
                print(f"[JOURNAL-PARSER] Found header at line {i}, starting parse at line {start_idx}")
                break

        # Process data
        current_transaction = None
        current_id = None

        for i in range(start_idx, len(lines)):
            line = lines[i].strip()

            # Skip empty lines and page footers/headers
            if not line or 'Accrual Basis' in line or ('Journal' in line and len(line) < 50):
                continue

            # Check for "Total for" line - marks end of transaction
            if line.startswith('Total for'):
                if current_transaction and current_transaction.get('lines'):
                    self.transactions.append(current_transaction)
                    print(f"[JOURNAL-PARSER] Completed transaction {current_transaction['id']} with {len(current_transaction['lines'])} lines")
                current_transaction = None
                current_id = None
                continue

            # Check if this is a transaction ID line (just a number)
            if line.isdigit() and len(line) <= 4:
                print(f"[JOURNAL-PARSER] Found transaction ID: {line}")
                if current_transaction and current_transaction.get('lines'):
                    self.transactions.append(current_transaction)
                current_id = line
                current_transaction = {
                    'id': current_id,
                    'lines': []
                }
                continue

            # Parse transaction detail line (pdfplumber single-space format)
            if current_transaction:
                # Extract amounts from end of line (1 or 2 decimal numbers)
                amounts = re.findall(r'[\d,]+\.\d{2}', line)
                if not amounts:
                    continue

                # Remove amounts from line
                line_without_amounts = line
                for amt in amounts:
                    line_without_amounts = line_without_amounts.rsplit(amt, 1)[0].strip()

                # Check for date at start
                date_match = re.match(r'^(\d{2}/\d{2}/\d{4})\s+(.+)$', line_without_amounts)

                if date_match:
                    # Line with date
                    date_str = date_match.group(1)
                    rest = date_match.group(2).strip()

                    # Set transaction header if not set
                    if not current_transaction.get('date'):
                        current_transaction['date'] = date_str
                        parts = rest.split()
                        current_transaction['type'] = parts[0] if len(parts) > 0 else ''
                        if len(parts) > 1 and parts[1].isdigit():
                            current_transaction['num'] = parts[1]
                        # Name is everything after type/num until account
                        if len(parts) > 2:
                            current_transaction['name'] = ' '.join(parts[2:] if not parts[1].isdigit() else parts[2:])

                    # Account is the text before amounts (typically last 3-5 words)
                    account_words = rest.split()[-5:]
                    account = ' '.join(account_words)

                    # Parse amounts
                    if len(amounts) == 2:
                        debit = float(amounts[0].replace(',', ''))
                        credit = float(amounts[1].replace(',', ''))
                    else:
                        debit = 0.0
                        credit = float(amounts[0].replace(',', ''))

                    if account and (debit > 0 or credit > 0):
                        current_transaction['lines'].append({
                            'account': account,
                            'description': '',
                            'debit': debit,
                            'credit': credit
                        })

                else:
                    # Continuation line (no date)
                    # Account is text before amounts
                    words = line_without_amounts.split()
                    account = ' '.join(words[-3:]) if len(words) >= 3 else line_without_amounts

                    # Parse amounts
                    if len(amounts) == 2:
                        debit = float(amounts[0].replace(',', ''))
                        credit = float(amounts[1].replace(',', ''))
                    else:
                        debit = 0.0
                        credit = float(amounts[0].replace(',', ''))

                    if account and (debit > 0 or credit > 0):
                        current_transaction['lines'].append({
                            'account': account,
                            'description': '',
                            'debit': debit,
                            'credit': credit
                        })

        # Don't forget the last transaction
        if current_transaction and current_transaction.get('lines'):
            self.transactions.append(current_transaction)

        print(f"[JOURNAL-PARSER] Parsed {len(self.transactions)} transactions from PDF")
        result = self.build_json_structure()
        print(f"[JOURNAL-PARSER] Built {len(result)} journal entries")
        return result

    def lookup_account_id(self, account_name):
        """Look up account ID by name"""
        if not self.use_account_lookup or not account_name:
            return None

        # Check cache first
        if account_name in self.account_cache:
            return self.account_cache[account_name]

        # Lookup via API
        account_id = self.lookup_client.lookup_account_id(account_name)
        if account_id:
            self.account_cache[account_name] = account_id

        return account_id

    def build_json_structure(self):
        """Build QuickBooks-compatible JSON structure from parsed transactions"""
        journal_entries = []

        for trans in self.transactions:
            if not trans.get('lines'):
                continue

            # Parse date
            date_obj = datetime.strptime(trans['date'], '%m/%d/%Y')

            # Filter for actual Journal Entry type transactions if needed
            # For now, we'll include all transaction types as they appear in the report

            entry = {
                "id": trans['id'],
                "syncToken": "0",
                "metaData": {
                    "createdByRef": None,
                    "createTime": date_obj.strftime('%Y-%m-%dT%H:%M:%S.000+00:00'),
                    "lastModifiedByRef": None,
                    "lastUpdatedTime": date_obj.strftime('%Y-%m-%dT%H:%M:%S.000+00:00'),
                    "lastChangedInQB": None,
                    "synchronized": None
                },
                "customField": [],
                "attachableRef": [],
                "domain": "QBO",
                "status": None,
                "sparse": False,
                "docNumber": trans.get('num') or None,
                "txnDate": date_obj.strftime('%Y-%m-%dT00:00:00.000+00:00'),
                "departmentRef": None,
                "currencyRef": {
                    "value": "USD",
                    "name": "United States Dollar",
                    "type": None
                },
                "exchangeRate": None,
                "privateNote": trans.get('memo', '') or None,
                "txnStatus": None,
                "linkedTxn": [],
                "line": [],
                "txnTaxDetail": {
                    "defaultTaxCodeRef": None,
                    "txnTaxCodeRef": None,
                    "totalTax": None,
                    "taxReviewStatus": None,
                    "taxLine": [],
                    "useAutomatedSalesTax": None
                },
                "txnSource": None,
                "taxFormType": None,
                "taxFormNum": None,
                "transactionLocationType": None,
                "tag": [],
                "txnApprovalInfo": None,
                "recurDataRef": None,
                "recurringInfo": None,
                "projectRef": None,
                "totalCostAmount": None,
                "homeTotalCostAmount": None,
                "adjustment": False,
                "homeCurrencyAdjustment": None,
                "enteredInHomeCurrency": None,
                "globalTaxCalculation": None,
                "totalAmt": 0.0,
                "homeTotalAmt": None,
                "journalEntryEx": None
            }

            # Process line items
            line_num = 0
            for line in trans['lines']:
                # Skip lines with no amounts
                if line['debit'] == 0 and line['credit'] == 0:
                    continue

                # Determine posting type and amount
                if line['debit'] > 0:
                    posting_type = "DEBIT"
                    amount = line['debit']
                else:
                    posting_type = "CREDIT"
                    amount = line['credit']

                # Look up account ID
                account_id = self.lookup_account_id(line['account'])

                line_item = {
                    "id": str(line_num),
                    "lineNum": None,
                    "description": line['description'] or None,
                    "amount": amount,
                    "received": None,
                    "linkedTxn": [],
                    "detailType": "JOURNAL_ENTRY_LINE_DETAIL",
                    "paymentLineDetail": None,
                    "discountLineDetail": None,
                    "taxLineDetail": None,
                    "salesItemLineDetail": None,
                    "descriptionLineDetail": None,
                    "itemBasedExpenseLineDetail": None,
                    "accountBasedExpenseLineDetail": None,
                    "reimburseLineDetail": None,
                    "depositLineDetail": None,
                    "purchaseOrderItemLineDetail": None,
                    "salesOrderItemLineDetail": None,
                    "itemReceiptLineDetail": None,
                    "journalEntryLineDetail": {
                        "postingType": posting_type,
                        "entity": None,
                        "accountRef": {
                            "value": account_id or str(100 + line_num),
                            "name": line['account'],
                            "type": None
                        },
                        "classRef": None,
                        "departmentRef": None,
                        "taxCodeRef": None,
                        "taxRateRef": None,
                        "taxApplicableOn": None,
                        "taxAmount": None,
                        "taxInclusiveAmt": None,
                        "billableStatus": None,
                        "journalCodeRef": None,
                        "journalEntryLineDetailEx": None
                    },
                    "groupLineDetail": None,
                    "subTotalLineDetail": None,
                    "itemAdjustmentLineDetail": None,
                    "customField": [],
                    "lineEx": None,
                    "projectRef": None,
                    "costAmount": None,
                    "homeCostAmount": None,
                    "tdslineDetail": None
                }

                entry['line'].append(line_item)
                line_num += 1

            # Calculate total amount (should be 0 for balanced entries)
            total_debits = sum(l['debit'] for l in trans['lines'])
            total_credits = sum(l['credit'] for l in trans['lines'])
            entry['totalAmt'] = round(abs(total_debits - total_credits), 2)

            journal_entries.append(entry)

        return journal_entries

    def save_to_file(self, journal_entries, output_file):
        """Save journal entries to JSON file"""
        with open(output_file, 'w') as f:
            json.dump(journal_entries, f, indent=2)


def main():
    import argparse

    parser = argparse.ArgumentParser(description='Convert journal entries to QuickBooks JSON format')
    parser.add_argument('input_file', help='Input file path (CSV, XLSX, or PDF)')
    parser.add_argument('-o', '--output', help='Output JSON file path')
    parser.add_argument('--no-lookup', action='store_true', help='Disable account lookup')
    parser.add_argument('--api-url', help='API base URL for account lookup')

    args = parser.parse_args()

    # Create converter
    converter = JournalEntriesConverter(
        use_account_lookup=not args.no_lookup,
        api_base_url=args.api_url
    )

    try:
        # Convert file
        entries = converter.convert(args.input_file)

        if args.output:
            # Save to file
            converter.save_to_file(entries, args.output)
            print(f"Successfully converted {len(entries)} journal entries to {args.output}")
        else:
            # Print to stdout
            print(json.dumps(entries, indent=2))

    except Exception as e:
        print(f"Error: {e}")
        return 1

    return 0


if __name__ == "__main__":
    exit(main())
