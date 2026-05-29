#!/usr/bin/env python3
"""
Accounts Receivable (A/R Aging Summary) Converter
Converts CSV, XLSX, and PDF A/R Aging Summary reports to QuickBooks JSON format
"""

import csv
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
import argparse
from typing import Dict, List, Any, Optional

from base_converter import BaseConverter, XLSX_SUPPORT, PDF_SUPPORT

if XLSX_SUPPORT:
    import openpyxl
if PDF_SUPPORT:
    import pdfplumber


class AccountsReceivableConverter(BaseConverter):
    """Converts A/R Aging Summary reports to QuickBooks-style JSON format"""

    def __init__(self, **kwargs):
        super().__init__(use_account_lookup=False)
        self.report_date = datetime.now(timezone.utc).strftime('%Y-%m-%d')
        self.customer_id = 1

    def create_header(self, report_date: str = None) -> Dict[str, Any]:
        """Create QuickBooks-style header for AgedReceivables report"""
        if report_date is None:
            report_date = self.report_date
        return self.make_qb_header(
            "AgedReceivables",
            start_period=report_date,
            end_period=report_date,
            options=[
                {"name": "report_date", "value": report_date},
                {"name": "NoReportData", "value": "false"}
            ]
        )

    def create_columns(self) -> Dict[str, List[Dict[str, Any]]]:
        """Create column definitions for A/R Aging report"""
        return {
            "column": [
                {"colTitle": "", "colType": "Customer", "metaData": [], "columns": None},
                {"colTitle": "Current", "colType": "Money", "metaData": [{"name": "ColKey", "value": "current"}], "columns": None},
                {"colTitle": "1 - 30", "colType": "Money", "metaData": [{"name": "ColKey", "value": "0"}], "columns": None},
                {"colTitle": "31 - 60", "colType": "Money", "metaData": [{"name": "ColKey", "value": "1"}], "columns": None},
                {"colTitle": "61 - 90", "colType": "Money", "metaData": [{"name": "ColKey", "value": "2"}], "columns": None},
                {"colTitle": "91 and over", "colType": "Money", "metaData": [{"name": "ColKey", "value": "3"}], "columns": None},
                {"colTitle": "Total", "colType": "Money", "metaData": [{"name": "ColKey", "value": "total"}], "columns": None},
            ]
        }

    def create_customer_row(self, customer_name: str, customer_id: str,
                            current: float, days_1_30: float, days_31_60: float,
                            days_61_90: float, days_91_over: float, total: float) -> Dict[str, Any]:
        """Create a customer row with aging data"""
        return {
            "id": None, "parentId": None, "header": None, "rows": None, "summary": None,
            "colData": [
                self.make_coldata_cell(customer_name, customer_id),
                self.make_coldata_cell(str(current) if current else ""),
                self.make_coldata_cell(str(days_1_30) if days_1_30 else ""),
                self.make_coldata_cell(str(days_31_60) if days_31_60 else ""),
                self.make_coldata_cell(str(days_61_90) if days_61_90 else ""),
                self.make_coldata_cell(str(days_91_over) if days_91_over else ""),
                self.make_coldata_cell(str(total)),
            ],
            "type": None, "group": None
        }

    def create_parent_customer_row(self, customer_name: str, customer_id: str,
                                    sub_customers: List[Dict[str, Any]],
                                    total_current: float, total_1_30: float, total_31_60: float,
                                    total_61_90: float, total_91_over: float, total: float) -> Dict[str, Any]:
        """Create a parent customer row with sub-customers"""
        return {
            "id": None, "parentId": None,
            "header": {
                "colData": [
                    self.make_coldata_cell(customer_name, customer_id),
                    self.make_coldata_cell(""),
                    self.make_coldata_cell(""),
                    self.make_coldata_cell(""),
                    self.make_coldata_cell(""),
                    self.make_coldata_cell(""),
                    self.make_coldata_cell("0.00"),
                ]
            },
            "rows": {"row": sub_customers},
            "summary": {
                "colData": [
                    self.make_coldata_cell(f"Total {customer_name}"),
                    self.make_coldata_cell(str(total_current)),
                    self.make_coldata_cell(str(total_1_30)),
                    self.make_coldata_cell(str(total_31_60)),
                    self.make_coldata_cell(str(total_61_90)),
                    self.make_coldata_cell(str(total_91_over)),
                    self.make_coldata_cell(str(total)),
                ]
            },
            "colData": [],
            "type": "SECTION", "group": None
        }

    def create_sub_customer_row(self, customer_name: str, customer_id: str,
                                current: float, days_1_30: float, days_31_60: float,
                                days_61_90: float, days_91_over: float, total: float) -> Dict[str, Any]:
        """Create a sub-customer row"""
        return {
            "id": None, "parentId": None, "header": None, "rows": None, "summary": None,
            "colData": [
                self.make_coldata_cell(customer_name, customer_id),
                self.make_coldata_cell(str(current) if current else ""),
                self.make_coldata_cell(str(days_1_30) if days_1_30 else ""),
                self.make_coldata_cell(str(days_31_60) if days_31_60 else ""),
                self.make_coldata_cell(str(days_61_90) if days_61_90 else ""),
                self.make_coldata_cell(str(days_91_over) if days_91_over else ""),
                self.make_coldata_cell(str(total)),
            ],
            "type": "DATA", "group": None
        }

    def create_total_row(self, current: float, days_1_30: float, days_31_60: float,
                         days_61_90: float, days_91_over: float, total: float) -> Dict[str, Any]:
        """Create the grand total summary row"""
        return {
            "id": None, "parentId": None, "header": None, "rows": None,
            "summary": {
                "colData": [
                    self.make_coldata_cell("TOTAL"),
                    self.make_coldata_cell(str(current)),
                    self.make_coldata_cell(str(days_1_30)),
                    self.make_coldata_cell(str(days_31_60)),
                    self.make_coldata_cell(str(days_61_90)),
                    self.make_coldata_cell(str(days_91_over)),
                    self.make_coldata_cell(str(total)),
                ]
            },
            "colData": [],
            "type": "SECTION", "group": "GrandTotal"
        }

    def parse_csv(self, filepath: Path) -> Dict[str, Any]:
        """Parse CSV file and convert to QuickBooks JSON format"""
        customers = []
        totals = {'current': 0.0, '1_30': 0.0, '31_60': 0.0, '61_90': 0.0, '91_over': 0.0, 'total': 0.0}

        with open(filepath, 'r', encoding='utf-8') as f:
            for _ in range(4):
                f.readline()

            reader = csv.DictReader(f)
            current_parent = None
            sub_customers = []
            parent_totals = {'current': 0.0, '1_30': 0.0, '31_60': 0.0, '61_90': 0.0, '91_over': 0.0, 'total': 0.0}

            for row in reader:
                customer_name = row.get('Customer', '').strip()

                if not customer_name:
                    continue

                if customer_name.upper() == 'TOTAL':
                    totals['current'] = self.parse_amount(row.get('CURRENT', '0'))
                    totals['1_30'] = self.parse_amount(row.get('1 - 30', '0'))
                    totals['31_60'] = self.parse_amount(row.get('31 - 60', '0'))
                    totals['61_90'] = self.parse_amount(row.get('61 - 90', '0'))
                    totals['91_over'] = self.parse_amount(row.get('91 AND OVER', '0'))
                    totals['total'] = self.parse_amount(row.get('Total', '0'))
                    break

                if customer_name.startswith('Total for '):
                    if current_parent and sub_customers:
                        parent_row = self.create_parent_customer_row(
                            current_parent, str(self.customer_id), sub_customers,
                            parent_totals['current'], parent_totals['1_30'], parent_totals['31_60'],
                            parent_totals['61_90'], parent_totals['91_over'], parent_totals['total']
                        )
                        customers.append(parent_row)
                        self.customer_id += 1
                        current_parent = None
                        sub_customers = []
                        parent_totals = {'current': 0.0, '1_30': 0.0, '31_60': 0.0, '61_90': 0.0, '91_over': 0.0, 'total': 0.0}
                    continue

                current = self.parse_amount(row.get('CURRENT', '0'))
                days_1_30 = self.parse_amount(row.get('1 - 30', '0'))
                days_31_60 = self.parse_amount(row.get('31 - 60', '0'))
                days_61_90 = self.parse_amount(row.get('61 - 90', '0'))
                days_91_over = self.parse_amount(row.get('91 AND OVER', '0'))
                total = self.parse_amount(row.get('Total', '0'))

                if current == 0 and days_1_30 == 0 and days_31_60 == 0 and days_61_90 == 0 and days_91_over == 0 and total == 0:
                    if current_parent and sub_customers:
                        parent_row = self.create_parent_customer_row(
                            current_parent, str(self.customer_id), sub_customers,
                            parent_totals['current'], parent_totals['1_30'], parent_totals['31_60'],
                            parent_totals['61_90'], parent_totals['91_over'], parent_totals['total']
                        )
                        customers.append(parent_row)
                        self.customer_id += 1
                        sub_customers = []
                        parent_totals = {'current': 0.0, '1_30': 0.0, '31_60': 0.0, '61_90': 0.0, '91_over': 0.0, 'total': 0.0}
                    current_parent = customer_name
                    continue

                if current_parent:
                    sub_row = self.create_sub_customer_row(
                        customer_name, str(self.customer_id),
                        current, days_1_30, days_31_60, days_61_90, days_91_over, total
                    )
                    sub_customers.append(sub_row)
                    self.customer_id += 1
                    parent_totals['current'] += current
                    parent_totals['1_30'] += days_1_30
                    parent_totals['31_60'] += days_31_60
                    parent_totals['61_90'] += days_61_90
                    parent_totals['91_over'] += days_91_over
                    parent_totals['total'] += total
                else:
                    customer_row = self.create_customer_row(
                        customer_name, str(self.customer_id),
                        current, days_1_30, days_31_60, days_61_90, days_91_over, total
                    )
                    customers.append(customer_row)
                    self.customer_id += 1

            if current_parent and sub_customers:
                parent_row = self.create_parent_customer_row(
                    current_parent, str(self.customer_id), sub_customers,
                    parent_totals['current'], parent_totals['1_30'], parent_totals['31_60'],
                    parent_totals['61_90'], parent_totals['91_over'], parent_totals['total']
                )
                customers.append(parent_row)
                self.customer_id += 1

        total_row = self.create_total_row(
            totals['current'], totals['1_30'], totals['31_60'],
            totals['61_90'], totals['91_over'], totals['total']
        )
        customers.append(total_row)

        return {
            "header": self.create_header(),
            "columns": self.create_columns(),
            "rows": {"row": customers}
        }

    def parse_xlsx(self, filepath: Path) -> Dict[str, Any]:
        """Parse XLSX file and convert to QuickBooks JSON format"""
        self.check_xlsx_support()

        customers = []
        totals = {'current': 0.0, '1_30': 0.0, '31_60': 0.0, '61_90': 0.0, '91_over': 0.0, 'total': 0.0}

        workbook = openpyxl.load_workbook(filepath, data_only=True)
        sheet = workbook.active

        header_row = None
        for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            # Look for aging bucket headers: Current, 1-30, 31-60, etc.
            row_str = ' '.join(str(cell) for cell in row if cell)
            if ('Current' in row_str and 'Total' in row_str) or \
               ('1' in row_str and '30' in row_str and '60' in row_str):
                header_row = idx
                break

        if not header_row:
            raise ValueError("Could not find header row in XLSX file")

        headers = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        col_map = {str(header).strip(): idx for idx, header in enumerate(headers) if header}

        current_parent = None
        sub_customers = []
        parent_totals = {'current': 0.0, '1_30': 0.0, '31_60': 0.0, '61_90': 0.0, '91_over': 0.0, 'total': 0.0}

        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[col_map.get('Customer', 0)]:
                continue

            customer_name = str(row[col_map.get('Customer', 0)]).strip()

            if customer_name.upper() == 'TOTAL':
                totals['current'] = self.parse_amount(str(row[col_map.get('CURRENT', 1)] or '0'))
                totals['1_30'] = self.parse_amount(str(row[col_map.get('1 - 30', 2)] or '0'))
                totals['31_60'] = self.parse_amount(str(row[col_map.get('31 - 60', 3)] or '0'))
                totals['61_90'] = self.parse_amount(str(row[col_map.get('61 - 90', 4)] or '0'))
                totals['91_over'] = self.parse_amount(str(row[col_map.get('91 AND OVER', 5)] or '0'))
                totals['total'] = self.parse_amount(str(row[col_map.get('Total', 6)] or '0'))
                break

            if customer_name.startswith('Total for '):
                if current_parent and sub_customers:
                    parent_row = self.create_parent_customer_row(
                        current_parent, str(self.customer_id), sub_customers,
                        parent_totals['current'], parent_totals['1_30'], parent_totals['31_60'],
                        parent_totals['61_90'], parent_totals['91_over'], parent_totals['total']
                    )
                    customers.append(parent_row)
                    self.customer_id += 1
                    current_parent = None
                    sub_customers = []
                    parent_totals = {'current': 0.0, '1_30': 0.0, '31_60': 0.0, '61_90': 0.0, '91_over': 0.0, 'total': 0.0}
                continue

            current = self.parse_amount(str(row[col_map.get('CURRENT', 1)] or '0'))
            days_1_30 = self.parse_amount(str(row[col_map.get('1 - 30', 2)] or '0'))
            days_31_60 = self.parse_amount(str(row[col_map.get('31 - 60', 3)] or '0'))
            days_61_90 = self.parse_amount(str(row[col_map.get('61 - 90', 4)] or '0'))
            days_91_over = self.parse_amount(str(row[col_map.get('91 AND OVER', 5)] or '0'))
            total = self.parse_amount(str(row[col_map.get('Total', 6)] or '0'))

            if current == 0 and days_1_30 == 0 and days_31_60 == 0 and days_61_90 == 0 and days_91_over == 0 and total == 0:
                if current_parent and sub_customers:
                    parent_row = self.create_parent_customer_row(
                        current_parent, str(self.customer_id), sub_customers,
                        parent_totals['current'], parent_totals['1_30'], parent_totals['31_60'],
                        parent_totals['61_90'], parent_totals['91_over'], parent_totals['total']
                    )
                    customers.append(parent_row)
                    self.customer_id += 1
                    sub_customers = []
                    parent_totals = {'current': 0.0, '1_30': 0.0, '31_60': 0.0, '61_90': 0.0, '91_over': 0.0, 'total': 0.0}
                current_parent = customer_name
                continue

            if current_parent:
                sub_row = self.create_sub_customer_row(
                    customer_name, str(self.customer_id),
                    current, days_1_30, days_31_60, days_61_90, days_91_over, total
                )
                sub_customers.append(sub_row)
                self.customer_id += 1
                parent_totals['current'] += current
                parent_totals['1_30'] += days_1_30
                parent_totals['31_60'] += days_31_60
                parent_totals['61_90'] += days_61_90
                parent_totals['91_over'] += days_91_over
                parent_totals['total'] += total
            else:
                customer_row = self.create_customer_row(
                    customer_name, str(self.customer_id),
                    current, days_1_30, days_31_60, days_61_90, days_91_over, total
                )
                customers.append(customer_row)
                self.customer_id += 1

        if current_parent and sub_customers:
            parent_row = self.create_parent_customer_row(
                current_parent, str(self.customer_id), sub_customers,
                parent_totals['current'], parent_totals['1_30'], parent_totals['31_60'],
                parent_totals['61_90'], parent_totals['91_over'], parent_totals['total']
            )
            customers.append(parent_row)
            self.customer_id += 1

        total_row = self.create_total_row(
            totals['current'], totals['1_30'], totals['31_60'],
            totals['61_90'], totals['91_over'], totals['total']
        )
        customers.append(total_row)

        return {
            "header": self.create_header(),
            "columns": self.create_columns(),
            "rows": {"row": customers}
        }

    # QB A/R Aging PDFs draw no cell borders. Reading the page with the table
    # extractor smears the bucket columns together (the header "CUSTOMER CURRENT
    # 1 - 30 ..." is split mid-word), so instead we read individual words with
    # their x-coordinates and assign each amount to a bucket by horizontal
    # position. This recovers every customer (including ones with several
    # bucket amounts) and handles names/labels that wrap onto a second line.
    _AMOUNT_RE = re.compile(r'-?\$?[\d,]+\.\d{2}')

    # Right edges (x1) of each aging-bucket column, derived from the report
    # header word positions. Amounts are right-aligned, so an amount belongs to
    # the bucket whose right edge its own right edge sits closest to.
    _BUCKET_EDGES = [
        (298, 'current'),
        (334, '1_30'),
        (378, '31_60'),
        (426, '61_90'),
        (507, '91_over'),
        (564, 'total'),
    ]

    @classmethod
    def _bucket_for_x(cls, x1: float) -> str:
        """Map an amount's right edge to its aging-bucket key."""
        return min(cls._BUCKET_EDGES, key=lambda e: abs(e[0] - x1))[1]

    def _extract_pdf_lines(self, filepath: Path) -> List[Dict[str, Any]]:
        """Read the PDF into logical report lines.

        Each line is ``{"name": str, "amounts": {bucket: value}}``. Words sharing
        the same vertical position are grouped; non-numeric words form the name,
        numeric words are placed into buckets by x-position. Report furniture
        (title, company name, "As of ...", page footer, the column header) is
        skipped. Only lines after the column header on each page are kept.
        """
        lines: List[Dict[str, Any]] = []
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                grouped: Dict[int, List[Dict[str, Any]]] = {}
                for w in page.extract_words():
                    grouped.setdefault(round(w['top']), []).append(w)

                seen_header = False
                for top in sorted(grouped):
                    words = sorted(grouped[top], key=lambda w: w['x0'])
                    name_parts: List[str] = []
                    amounts: Dict[str, float] = {}
                    for w in words:
                        text = w['text']
                        if self._AMOUNT_RE.fullmatch(text):
                            bucket = self._bucket_for_x(w['x1'])
                            amounts[bucket] = self.parse_amount(text)
                        else:
                            name_parts.append(text)
                    name = ' '.join(name_parts).strip()

                    # Locate / skip the column header ("CUSTOMER ... TOTAL").
                    if not seen_header:
                        upper = name.upper()
                        if upper.startswith('CUSTOMER') and ('CURRENT' in upper or 'TOTAL' in upper):
                            seen_header = True
                        continue

                    # Skip page footer (e.g. "Wednesday, March 04, 2026 ... GMTZ 2/2").
                    if 'GMTZ' in name or re.search(r'\b\d+/\d+\b', name):
                        continue

                    if not name and not amounts:
                        continue

                    lines.append({'name': name, 'amounts': amounts})
        return lines

    def parse_pdf(self, filepath: Path) -> Dict[str, Any]:
        """Parse PDF file and convert to QuickBooks JSON format.

        Mirrors the parent/sub-customer structure produced by parse_csv /
        parse_xlsx so downstream consumers see an identical envelope.
        """
        self.check_pdf_support()

        raw_lines = self._extract_pdf_lines(filepath)

        # Merge name-wrap lines: an amount-less line that is NOT a group header
        # (i.e. not followed by sub-rows ending in "Total for") is the wrapped
        # tail of the line above it.
        def is_group_header(idx: int) -> bool:
            for j in range(idx + 1, len(raw_lines)):
                nxt = raw_lines[j]['name']
                if nxt.startswith('Total for'):
                    return True
                if not raw_lines[j]['amounts']:
                    # Another amount-less line before any "Total for" -> not a header.
                    return False
            return False

        merged: List[Dict[str, Any]] = []
        for i, line in enumerate(raw_lines):
            name, amounts = line['name'], line['amounts']
            if (not amounts and name and merged
                    and not name.startswith('Total for')
                    and name.upper() != 'TOTAL'
                    and not is_group_header(i)
                    and not merged[-1]['name'].startswith('Total for')
                    and merged[-1]['name'].upper() != 'TOTAL'):
                # Only merge into the previous line if that line is itself an
                # amount-less group header awaiting its name tail, OR a plain
                # data/customer line whose name wrapped. We never merge into a
                # "Total for" / "TOTAL" line.
                merged[-1]['name'] = f"{merged[-1]['name']} {name}".strip()
                # If the previous line had amounts it was a data row whose name
                # wrapped; keep its amounts. If it was amount-less it stays a
                # header candidate. Re-evaluation of headers happens below.
                continue
            merged.append({'name': name, 'amounts': dict(amounts)})

        customers: List[Dict[str, Any]] = []
        totals = {'current': 0.0, '1_30': 0.0, '31_60': 0.0, '61_90': 0.0, '91_over': 0.0, 'total': 0.0}
        current_parent: Optional[str] = None
        sub_customers: List[Dict[str, Any]] = []
        parent_totals = {'current': 0.0, '1_30': 0.0, '31_60': 0.0, '61_90': 0.0, '91_over': 0.0, 'total': 0.0}

        def flush_parent():
            nonlocal current_parent, sub_customers, parent_totals
            if current_parent and sub_customers:
                parent_row = self.create_parent_customer_row(
                    current_parent, str(self.customer_id), sub_customers,
                    parent_totals['current'], parent_totals['1_30'], parent_totals['31_60'],
                    parent_totals['61_90'], parent_totals['91_over'], parent_totals['total']
                )
                customers.append(parent_row)
                self.customer_id += 1
            current_parent = None
            sub_customers = []
            parent_totals = {'current': 0.0, '1_30': 0.0, '31_60': 0.0, '61_90': 0.0, '91_over': 0.0, 'total': 0.0}

        for line in merged:
            name = line['name']
            amounts = line['amounts']

            if not name:
                continue

            if name.upper() == 'TOTAL':
                totals['current'] = amounts.get('current', 0.0)
                totals['1_30'] = amounts.get('1_30', 0.0)
                totals['31_60'] = amounts.get('31_60', 0.0)
                totals['61_90'] = amounts.get('61_90', 0.0)
                totals['91_over'] = amounts.get('91_over', 0.0)
                totals['total'] = amounts.get('total', 0.0)
                break

            if name.startswith('Total for '):
                flush_parent()
                continue

            current = amounts.get('current', 0.0)
            days_1_30 = amounts.get('1_30', 0.0)
            days_31_60 = amounts.get('31_60', 0.0)
            days_61_90 = amounts.get('61_90', 0.0)
            days_91_over = amounts.get('91_over', 0.0)
            total = amounts.get('total', 0.0)

            # Amount-less line = a group/parent header (its children follow).
            if not amounts:
                flush_parent()
                current_parent = name
                continue

            if current_parent:
                sub_row = self.create_sub_customer_row(
                    name, str(self.customer_id),
                    current, days_1_30, days_31_60, days_61_90, days_91_over, total
                )
                sub_customers.append(sub_row)
                self.customer_id += 1
                parent_totals['current'] += current
                parent_totals['1_30'] += days_1_30
                parent_totals['31_60'] += days_31_60
                parent_totals['61_90'] += days_61_90
                parent_totals['91_over'] += days_91_over
                parent_totals['total'] += total
            else:
                customer_row = self.create_customer_row(
                    name, str(self.customer_id),
                    current, days_1_30, days_31_60, days_61_90, days_91_over, total
                )
                customers.append(customer_row)
                self.customer_id += 1

        flush_parent()

        print(f"[AR-PARSER] Parsed {len(customers)} customer rows")

        total_row = self.create_total_row(
            totals['current'], totals['1_30'], totals['31_60'],
            totals['61_90'], totals['91_over'], totals['total']
        )
        customers.append(total_row)

        return {
            "header": self.create_header(),
            "columns": self.create_columns(),
            "rows": {"row": customers}
        }

    def convert_to_json(self, filepath: Path, output_path: Optional[Path] = None) -> str:
        """Convert a file to JSON format"""
        import json
        result = self.convert_file(filepath)

        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, indent=2)
            customer_count = len([r for r in result['rows']['row'] if r.get('type') != 'SECTION' or r.get('group') != 'GrandTotal'])
            return f"Converted {customer_count} customers to {output_path}"
        else:
            return json.dumps(result, indent=2)


def main():
    parser = argparse.ArgumentParser(description='Convert A/R Aging Summary reports to JSON format')
    parser.add_argument('input', help='Input file (CSV, XLSX, or PDF)')
    parser.add_argument('-o', '--output', help='Output JSON file (default: print to stdout)')
    parser.add_argument('--batch', action='store_true', help='Process all files in a directory')

    args = parser.parse_args()

    converter = AccountsReceivableConverter()

    if args.batch:
        input_path = Path(args.input)
        if not input_path.is_dir():
            print(f"Error: {input_path} is not a directory", file=sys.stderr)
            sys.exit(1)

        output_dir = Path(args.output) if args.output else input_path.parent / 'converted'
        output_dir.mkdir(exist_ok=True)

        for file in input_path.glob('*'):
            if file.suffix.lower() in ['.csv', '.xlsx', '.pdf']:
                try:
                    output_file = output_dir / f"{file.stem}_ar_aging.json"
                    result = converter.convert_to_json(file, output_file)
                    print(result)
                except Exception as e:
                    print(f"Error processing {file}: {e}", file=sys.stderr)
    else:
        input_path = Path(args.input)
        if not input_path.exists():
            print(f"Error: {input_path} does not exist", file=sys.stderr)
            sys.exit(1)

        try:
            if args.output:
                result = converter.convert_to_json(input_path, Path(args.output))
                print(result)
            else:
                print(converter.convert_to_json(input_path))
        except Exception as e:
            print(f"Error: {e}", file=sys.stderr)
            sys.exit(1)


if __name__ == '__main__':
    main()
