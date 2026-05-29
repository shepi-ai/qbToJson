#!/usr/bin/env python3
"""
Chart of Accounts Document Converter
Converts CSV, XLSX, and PDF account lists to QuickBooks JSON format
Specifically designed for Chart of Accounts reports
"""

import csv
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
import argparse
from typing import Dict, List, Any, Optional

from base_converter import BaseConverter, XLSX_SUPPORT, PDF_SUPPORT

if XLSX_SUPPORT:
    import openpyxl
if PDF_SUPPORT:
    import pdfplumber


class AccountsConverter(BaseConverter):
    """Converts Chart of Accounts documents to QuickBooks-style JSON format"""

    def __init__(self, **kwargs):
        super().__init__(use_account_lookup=False)
        self.account_id_counter = 200  # Starting ID for generated accounts

    def get_classification_from_type(self, type_str: str) -> str:
        """Determine classification based on type string"""
        type_lower = type_str.lower()
        if 'accounts payable' in type_lower or type_lower == 'a/p':
            return 'LIABILITY'
        elif 'accounts receivable' in type_lower or type_lower == 'a/r':
            return 'ASSET'
        elif type_lower == 'bank' or type_lower == 'checking' or type_lower == 'savings':
            return 'ASSET'
        elif 'credit card' in type_lower:
            return 'LIABILITY'
        elif 'cost of goods sold' in type_lower or type_lower == 'cogs':
            return 'EXPENSE'
        elif 'equity' in type_lower:
            return 'EQUITY'
        elif 'expense' in type_lower:
            return 'EXPENSE'
        elif 'income' in type_lower:
            return 'REVENUE'
        elif 'asset' in type_lower:
            return 'ASSET'
        elif 'liabilit' in type_lower:
            return 'LIABILITY'
        else:
            return 'EXPENSE'

    def get_account_type_from_type(self, type_str: str) -> str:
        """Determine account type based on type string"""
        type_lower = type_str.lower()
        if 'accounts payable' in type_lower or type_lower == 'a/p':
            return 'ACCOUNTS_PAYABLE'
        elif 'accounts receivable' in type_lower or type_lower == 'a/r':
            return 'ACCOUNTS_RECEIVABLE'
        elif type_lower == 'bank' or type_lower == 'checking' or type_lower == 'savings':
            return 'BANK'
        elif 'credit card' in type_lower:
            return 'CREDIT_CARD'
        elif 'cost of goods sold' in type_lower or type_lower == 'cogs':
            return 'COST_OF_GOODS_SOLD'
        elif 'equity' in type_lower:
            return 'EQUITY'
        elif 'other expense' in type_lower:
            return 'OTHER_EXPENSE'
        elif 'expense' in type_lower:
            return 'EXPENSE'
        elif 'other income' in type_lower:
            return 'OTHER_INCOME'
        elif 'income' in type_lower:
            return 'INCOME'
        elif 'other current asset' in type_lower:
            return 'OTHER_CURRENT_ASSET'
        elif 'fixed asset' in type_lower:
            return 'FIXED_ASSET'
        elif 'asset' in type_lower:
            return 'OTHER_CURRENT_ASSET'
        elif 'other current liabilit' in type_lower:
            return 'OTHER_CURRENT_LIABILITY'
        elif 'current liabilit' in type_lower:
            return 'CURRENT_LIABILITY'
        elif 'long term liabilit' in type_lower:
            return 'LONG_TERM_LIABILITY'
        elif 'liabilit' in type_lower:
            return 'OTHER_CURRENT_LIABILITY'
        else:
            return 'EXPENSE'

    # QuickBooks emits a fixed, closed set of account types, whereas account
    # *names* are freeform. So we recognise real accounts by their Type, not by
    # blocklisting names: report decoration (title, company name, the TOTAL row,
    # the "Accrual Basis ... GMTZ" footer) has an empty Type, while every real
    # account -- however oddly named -- has one of these.
    ACCOUNT_TYPE_TOKENS = (
        'accounts payable', 'a/p', 'accounts receivable', 'a/r', 'bank',
        'checking', 'savings', 'credit card', 'cost of goods sold', 'cogs',
        'equity', 'expense', 'income', 'asset', 'liabilit',
    )

    def is_known_account_type(self, type_str: str) -> bool:
        """True if the Type cell names a real QuickBooks account type."""
        t = (type_str or '').strip().lower()
        return bool(t) and any(tok in t for tok in self.ACCOUNT_TYPE_TOKENS)

    def create_account_object(self, name: str, type_str: str, detail_type: str,
                              description: Optional[str] = None,
                              balance: float = 0.0,
                              parent_id: Optional[str] = None) -> Dict[str, Any]:
        """Create a QuickBooks-style account object"""
        classification = self.get_classification_from_type(type_str)
        account_type = self.get_account_type_from_type(type_str)

        account_subtype = detail_type.strip()
        account_subtype = account_subtype.replace('/', '')
        account_subtype = account_subtype.replace('&', 'And')
        account_subtype = account_subtype.replace(' ', '')

        timestamp = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.000+00:00')

        # Detect sub-account from colon-separated name
        is_sub_account = ':' in name
        if is_sub_account:
            fully_qualified_name = name
            leaf_name = name.split(':')[-1].strip()
            parent_ref = {"value": parent_id, "name": None, "type": None} if parent_id else None
        else:
            fully_qualified_name = name
            leaf_name = name
            parent_ref = None

        return {
            "id": self.generate_account_id(),
            "syncToken": "0",
            "metaData": {
                "createdByRef": None,
                "createTime": timestamp,
                "lastModifiedByRef": None,
                "lastUpdatedTime": timestamp,
                "lastChangedInQB": None,
                "synchronized": None
            },
            "customField": [],
            "attachableRef": [],
            "domain": "QBO",
            "status": None,
            "sparse": False,
            "name": leaf_name,
            "subAccount": is_sub_account,
            "parentRef": parent_ref,
            "description": description if description else None,
            "fullyQualifiedName": fully_qualified_name,
            "accountAlias": None,
            "txnLocationType": None,
            "active": True,
            "classification": classification,
            "accountType": account_type,
            "accountSubType": account_subtype,
            "accountPurposes": [],
            "acctNum": None,
            "acctNumExtn": None,
            "bankNum": None,
            "openingBalance": None,
            "openingBalanceDate": None,
            "currentBalance": balance,
            "currentBalanceWithSubAccounts": balance,
            "currencyRef": {
                "value": "USD",
                "name": "United States Dollar",
                "type": None
            },
            "taxAccount": None,
            "taxCodeRef": None,
            "onlineBankingEnabled": None,
            "journalCodeRef": None,
            "accountEx": None,
            "finame": None
        }

    def parse_csv(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse CSV file and convert to account objects"""
        accounts = []
        # Track parent account IDs for sub-account references
        parent_ids = {}  # Maps parent name to its generated ID

        with open(filepath, 'r', encoding='utf-8') as f:
            all_lines = f.readlines()

        print(f"[CSV Parser] Total lines in file: {len(all_lines)}")

        header_line_idx = -1
        for idx, line in enumerate(all_lines):
            line_lower = line.lower()
            if 'full name' in line_lower or ('name' in line_lower and 'type' in line_lower):
                header_line_idx = idx
                print(f"[CSV Parser] Found header at line {idx}: {line.strip()}")
                break

        if header_line_idx == -1:
            print(f"[CSV Parser] WARNING: No header row found. First 5 lines:")
            for i, line in enumerate(all_lines[:5]):
                print(f"  Line {i}: {line.strip()}")
            return accounts

        from io import StringIO
        csv_content = ''.join(all_lines[header_line_idx:])
        reader = csv.DictReader(StringIO(csv_content))

        print(f"[CSV Parser] Column headers: {reader.fieldnames}")

        row_count = 0
        skipped_count = 0

        for row in reader:
            row_count += 1

            full_name = (row.get('Full name') or row.get('Name') or
                         row.get('FULL NAME') or row.get('NAME') or '').strip()

            if not full_name:
                skipped_count += 1
                continue

            type_str = (row.get('Type') or row.get('TYPE') or
                        row.get('Account Type') or '').strip()

            # Keep only real accounts (identified by a recognised Type). Report
            # decoration -- the TOTAL row and the "Accrual Basis ... GMTZ" footer --
            # has no Type and is skipped, without blocklisting freeform names.
            if not self.is_known_account_type(type_str):
                skipped_count += 1
                continue

            detail_type = (row.get('Detail type') or row.get('Detail Type') or
                           row.get('DETAIL TYPE') or row.get('Sub Type') or '').strip()
            description = (row.get('Description') or row.get('DESCRIPTION') or '').strip()

            balance_str = (row.get('Total balance') or row.get('Balance') or
                           row.get('TOTAL BALANCE') or row.get('Current Balance') or '0')
            balance_str = str(balance_str).replace('$', '').replace(',', '').strip()
            try:
                balance = float(balance_str) if balance_str else 0.0
            except ValueError:
                balance = 0.0

            # Determine parent ID for sub-accounts
            parent_id = None
            if ':' in full_name:
                parent_name = full_name.rsplit(':', 1)[0]
                parent_id = parent_ids.get(parent_name)

            account = self.create_account_object(
                name=full_name,
                type_str=type_str,
                detail_type=detail_type or 'Other',
                description=description,
                balance=balance,
                parent_id=parent_id
            )
            accounts.append(account)

            # Track this account's ID for potential sub-accounts
            parent_ids[full_name] = account['id']

        print(f"[CSV Parser] Processed {row_count} rows, created {len(accounts)} accounts, skipped {skipped_count} rows")

        return accounts

    def parse_xlsx(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse XLSX file and convert to account objects"""
        self.check_xlsx_support()

        accounts = []
        parent_ids = {}
        workbook = openpyxl.load_workbook(filepath, data_only=True)
        sheet = workbook.active

        header_row = None
        for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row and 'Full name' in str(row):
                header_row = idx
                break

        if not header_row:
            raise ValueError("Could not find header row in XLSX file")

        headers = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        col_map = {header: idx for idx, header in enumerate(headers) if header}

        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[col_map.get('Full name', 0)]:
                continue

            name = str(row[col_map.get('Full name', 0)])
            type_str = str(row[col_map.get('Type', 1)] or '').strip()

            # Keep only real accounts (recognised Type); the TOTAL row and the
            # "Accrual Basis ... GMTZ" footer have no Type and are skipped.
            if not self.is_known_account_type(type_str):
                continue

            balance_str = str(row[col_map.get('Total balance', 4)] or '0')
            balance_str = balance_str.replace('$', '').replace(',', '')
            try:
                balance = float(balance_str) if balance_str else 0.0
            except ValueError:
                balance = 0.0

            # Determine parent ID for sub-accounts
            parent_id = None
            if ':' in name:
                parent_name = name.rsplit(':', 1)[0]
                parent_id = parent_ids.get(parent_name)

            account = self.create_account_object(
                name=name,
                type_str=type_str,
                detail_type=row[col_map.get('Detail type', 2)] or '',
                description=row[col_map.get('Description', 3)],
                balance=balance,
                parent_id=parent_id
            )
            accounts.append(account)
            parent_ids[name] = account['id']

        return accounts

    # Table-style extraction settings (QB PDFs draw no cell borders). Retained
    # for reference; the word-position approach below proved more reliable for
    # the Account List layout where long names/types wrap onto extra lines.
    _PDF_TABLE_SETTINGS = {"vertical_strategy": "text", "horizontal_strategy": "text"}

    # Words at or below this top-coordinate gap are treated as the same visual row.
    _ROW_TOLERANCE = 3

    @staticmethod
    def _group_words_into_rows(words: List[Dict[str, Any]], tol: float) -> List[List[Dict[str, Any]]]:
        """Group extracted words into visual rows by their vertical position."""
        rows: List[List[Dict[str, Any]]] = []
        for w in sorted(words, key=lambda w: (round(w['top']), w['x0'])):
            if rows and abs(w['top'] - rows[-1][0]['top']) <= tol:
                rows[-1].append(w)
            else:
                rows.append([w])
        for r in rows:
            r.sort(key=lambda w: w['x0'])
        return rows

    @staticmethod
    def _is_report_furniture(text: str) -> bool:
        """True for QB title/footer rows that are not account data."""
        low = text.lower()
        if not low:
            return True
        if 'gmtz' in low or 'accrual basis' in low or 'cash basis' in low:
            return True
        if 'account list' in low or 'sandbox company' in low:
            return True
        # Footer date line, e.g. "Wednesday, March 04, 2026 10:53 PM GMTZ 1/5"
        if re.search(r'\b\d{1,2}:\d{2}\b', low) and re.search(r'\d/\d', low):
            return True
        return False

    def parse_pdf(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse a QuickBooks "Account List" PDF into account objects.

        QB renders these reports border-less and wraps long FULL NAME / TYPE /
        DETAIL TYPE values onto following lines that the text extractor returns
        as separate, partial rows. We recover columns from each word's x-position
        (using the per-page header to locate column boundaries, since they shift
        page to page) and merge wrapped continuation rows back into the account
        row above them. Output matches parse_csv / parse_xlsx exactly.
        """
        self.check_pdf_support()

        accounts: List[Dict[str, Any]] = []
        parent_ids: Dict[str, str] = {}

        # A pending account being assembled across one or more wrapped rows.
        pending: Optional[Dict[str, Any]] = None

        def flush():
            """Materialise the pending account into the accounts list."""
            nonlocal pending
            if pending is None:
                return
            name = pending['name'].strip()
            # Keep only real accounts (recognised Type), matching parse_csv /
            # parse_xlsx. Report decoration (title, company, TOTAL, the
            # "Accrual Basis ... GMTZ" footer) has no Type and is dropped.
            if name and self.is_known_account_type(pending['type']):
                parent_id = None
                if ':' in name:
                    parent_name = name.rsplit(':', 1)[0]
                    parent_id = parent_ids.get(parent_name)
                account = self.create_account_object(
                    name=name,
                    type_str=pending['type'].strip(),
                    detail_type=pending['detail'].strip() or 'Other',
                    description=pending['description'].strip() or None,
                    balance=pending['balance'],
                    parent_id=parent_id,
                )
                accounts.append(account)
                parent_ids[name] = account['id']
            pending = None

        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                words = page.extract_words(use_text_flow=False)
                if not words:
                    continue

                rows = self._group_words_into_rows(words, self._ROW_TOLERANCE)

                # Locate the header row and derive column x-boundaries from it.
                type_x = detail_x = desc_x = balance_x = None
                data_start = None
                for ridx, row in enumerate(rows):
                    texts = [w['text'] for w in row]
                    joined = ' '.join(texts)
                    if 'FULL' in texts and 'NAME' in texts and 'DETAIL' in texts:
                        # Standalone "TYPE" is the first TYPE not preceded by DETAIL.
                        for i, w in enumerate(row):
                            if w['text'] == 'TYPE':
                                prev = row[i - 1]['text'] if i > 0 else ''
                                if prev != 'DETAIL' and type_x is None:
                                    type_x = w['x0']
                            elif w['text'] == 'DETAIL':
                                detail_x = w['x0']
                            elif w['text'] == 'DESCRIPTION':
                                desc_x = w['x0']
                            elif w['text'] == 'TOTAL':
                                balance_x = w['x0']
                        data_start = ridx + 1
                        break

                if data_start is None or type_x is None or detail_x is None:
                    continue

                # Column thresholds. DETAIL TYPE and DESCRIPTION share a region in
                # this report (DESCRIPTION is empty); treat everything between the
                # DETAIL boundary and the balance column as DETAIL TYPE.
                balance_boundary = balance_x if balance_x is not None else 1e9

                for row in rows[data_start:]:
                    joined = ' '.join(w['text'] for w in row).strip()
                    if not joined:
                        continue
                    if self._is_report_furniture(joined):
                        # A trailing footer ends the page's data; flush pending.
                        continue

                    name_parts, type_parts, detail_parts, bal_parts = [], [], [], []
                    for w in row:
                        x = w['x0']
                        text = w['text']
                        # Balance values right-align, so their x-position varies a
                        # lot; identify them by content (currency / number) and the
                        # fact that they sit in the right-hand region of the page.
                        is_amount = bool(re.fullmatch(r'-?\$?[\d,]+\.?\d*', text))
                        if is_amount and x >= detail_x:
                            bal_parts.append(text)
                        elif x < type_x - 2:
                            name_parts.append(text)
                        elif x < detail_x - 2:
                            type_parts.append(text)
                        else:
                            detail_parts.append(text)

                    name_text = ' '.join(name_parts).strip()
                    type_text = ' '.join(type_parts).strip()
                    detail_text = ' '.join(detail_parts).strip()
                    bal_text = ' '.join(bal_parts).strip()

                    # The grand-total line: name == TOTAL with a balance, no type.
                    if name_text.upper() == 'TOTAL' and not type_text:
                        flush()
                        continue

                    # A new account row carries BOTH a FULL NAME and a TYPE value.
                    # Rows missing either are wrapped continuations of the account
                    # above (a wrapped name has no type; a wrapped multi-word
                    # type/detail has no name).
                    if name_text and type_text:
                        flush()
                        pending = {
                            'name': name_text,
                            'type': type_text,
                            'detail': detail_text,
                            'description': '',
                            'balance': self.parse_amount(bal_text) if bal_text else 0.0,
                        }
                    elif pending is not None:
                        # Continuation: append each fragment to its column.
                        if name_text:
                            pending['name'] = (pending['name'] + ' ' + name_text).strip()
                        if type_text:
                            pending['type'] = (pending['type'] + ' ' + type_text).strip()
                        if detail_text:
                            pending['detail'] = (pending['detail'] + ' ' + detail_text).strip()
                        if bal_text and pending['balance'] == 0.0:
                            pending['balance'] = self.parse_amount(bal_text)
                    # else: stray fragment before any account; ignore.

                flush()  # end of page

        return accounts


def main():
    parser = argparse.ArgumentParser(description='Convert account documents to JSON format')
    parser.add_argument('input', help='Input file (CSV, XLSX, or PDF)')
    parser.add_argument('-o', '--output', help='Output JSON file (default: print to stdout)')
    parser.add_argument('--batch', action='store_true', help='Process all files in a directory')

    args = parser.parse_args()

    converter = AccountsConverter()

    if args.batch:
        input_path = Path(args.input)
        if not input_path.is_dir():
            print(f"Error: {input_path} is not a directory", file=sys.stderr)
            sys.exit(1)

        output_dir = Path(args.output) if args.output else input_path.parent / 'converted'
        output_dir.mkdir(exist_ok=True)

        for file in input_path.glob('*'):
            if file.suffix.lower() in ['.csv', '.xlsx', '.pdf']:
                try:
                    output_file = output_dir / f"{file.stem}.json"
                    result = converter.convert_to_json(file, output_file)
                    print(result)
                except Exception as e:
                    print(f"Error processing {file}: {e}", file=sys.stderr)
    else:
        input_path = Path(args.input)
        if not input_path.exists():
            print(f"Error: {input_path} does not exist", file=sys.stderr)
            sys.exit(1)

        try:
            if args.output:
                result = converter.convert_to_json(input_path, Path(args.output))
                print(result)
            else:
                print(converter.convert_to_json(input_path))
        except Exception as e:
            print(f"Error: {e}", file=sys.stderr)
            sys.exit(1)


if __name__ == '__main__':
    main()
