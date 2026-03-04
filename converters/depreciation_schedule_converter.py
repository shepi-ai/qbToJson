#!/usr/bin/env python3
"""
Depreciation Schedule Converter
Converts CSV, XLSX, and PDF depreciation schedule reports to JSON array format.

Output format: flat array of depreciation line items per asset per period.
"""

import csv
import re
import sys
from pathlib import Path
import argparse
from typing import Dict, List, Any, Optional

from base_converter import BaseConverter, XLSX_SUPPORT, PDF_SUPPORT

if XLSX_SUPPORT:
    import openpyxl
if PDF_SUPPORT:
    import pdfplumber


class DepreciationScheduleConverter(BaseConverter):
    """Converts Depreciation Schedule reports to JSON array format."""

    def __init__(self, **kwargs):
        super().__init__(use_account_lookup=False)

    def parse_csv(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse CSV depreciation schedule."""
        results = []

        with open(filepath, 'r', encoding='utf-8') as f:
            all_lines = f.readlines()

        # Find header row
        header_line_idx = -1
        for idx, line in enumerate(all_lines):
            line_lower = line.lower()
            if 'asset' in line_lower and ('depreciation' in line_lower or 'expense' in line_lower):
                header_line_idx = idx
                break

        if header_line_idx == -1:
            # Try simpler header detection
            for idx, line in enumerate(all_lines):
                line_lower = line.lower()
                if 'asset' in line_lower and ('period' in line_lower or 'date' in line_lower or 'beginning' in line_lower):
                    header_line_idx = idx
                    break

        if header_line_idx == -1:
            print("[DEPRECIATION-PARSER] WARNING: No header row found")
            return results

        from io import StringIO
        csv_content = ''.join(all_lines[header_line_idx:])
        reader = csv.DictReader(StringIO(csv_content))

        print(f"[DEPRECIATION-PARSER] Column headers: {reader.fieldnames}")

        for row in reader:
            asset_name = (row.get('Asset') or row.get('Asset Name') or
                          row.get('ASSET') or row.get('Description') or '').strip()

            if not asset_name or asset_name.upper() == 'TOTAL':
                continue

            # Skip metadata rows
            if any(kw in asset_name.lower() for kw in ['total', 'grand total', 'summary']):
                continue

            period = (row.get('Period') or row.get('Date') or row.get('Month') or
                      row.get('PERIOD') or '').strip()

            beginning = self.parse_amount(
                row.get('Beginning Balance') or row.get('Beginning') or
                row.get('Opening Balance') or row.get('BEGINNING BALANCE') or '0'
            )
            expense = self.parse_amount(
                row.get('Depreciation Expense') or row.get('Depreciation') or
                row.get('Current Period') or row.get('DEPRECIATION') or '0'
            )
            ending = self.parse_amount(
                row.get('Ending Balance') or row.get('Ending') or
                row.get('Closing Balance') or row.get('ENDING BALANCE') or '0'
            )

            category = (row.get('Category') or row.get('Asset Class') or
                         row.get('Type') or '').strip()
            method = (row.get('Method') or row.get('Depreciation Method') or '').strip()

            results.append({
                "assetName": asset_name,
                "assetId": None,
                "period": period,
                "beginningBalance": beginning,
                "depreciationExpense": expense,
                "endingBalance": ending,
                "category": category,
                "depreciationMethod": method,
            })

        print(f"[DEPRECIATION-PARSER] Parsed {len(results)} depreciation lines")
        return results

    def parse_xlsx(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse XLSX depreciation schedule."""
        self.check_xlsx_support()

        results = []

        workbook = openpyxl.load_workbook(filepath, data_only=True)
        sheet = workbook.active

        # Find header row
        header_row = None
        for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row and any(cell and 'asset' in str(cell).lower() for cell in row):
                if any(cell and ('depreciation' in str(cell).lower() or 'expense' in str(cell).lower() or 'period' in str(cell).lower()) for cell in row):
                    header_row = idx
                    break

        if not header_row:
            raise ValueError("Could not find header row in XLSX file")

        headers = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        col_map = {str(header).strip(): idx for idx, header in enumerate(headers) if header}

        # Find relevant column indices
        asset_col = col_map.get('Asset') or col_map.get('Asset Name') or col_map.get('Description', 0)
        period_col = col_map.get('Period') or col_map.get('Date') or col_map.get('Month')
        begin_col = col_map.get('Beginning Balance') or col_map.get('Beginning') or col_map.get('Opening Balance')
        expense_col = col_map.get('Depreciation Expense') or col_map.get('Depreciation') or col_map.get('Current Period')
        ending_col = col_map.get('Ending Balance') or col_map.get('Ending') or col_map.get('Closing Balance')
        category_col = col_map.get('Category') or col_map.get('Asset Class') or col_map.get('Type')
        method_col = col_map.get('Method') or col_map.get('Depreciation Method')

        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[asset_col]:
                continue

            asset_name = str(row[asset_col]).strip()
            if asset_name.upper() in ('TOTAL', 'GRAND TOTAL'):
                continue

            results.append({
                "assetName": asset_name,
                "assetId": None,
                "period": str(row[period_col] or '') if period_col is not None else '',
                "beginningBalance": self.parse_amount(row[begin_col]) if begin_col is not None else 0.0,
                "depreciationExpense": self.parse_amount(row[expense_col]) if expense_col is not None else 0.0,
                "endingBalance": self.parse_amount(row[ending_col]) if ending_col is not None else 0.0,
                "category": str(row[category_col] or '') if category_col is not None else '',
                "depreciationMethod": str(row[method_col] or '') if method_col is not None else '',
            })

        print(f"[DEPRECIATION-PARSER] Parsed {len(results)} depreciation lines from XLSX")
        return results

    def parse_pdf(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse PDF depreciation schedule."""
        self.check_pdf_support()

        results = []

        print("[DEPRECIATION-PARSER] Starting PDF parse")

        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue

                lines = text.split('\n')

                # Find header
                header_idx = -1
                for i, line in enumerate(lines):
                    if 'ASSET' in line.upper() and ('DEPRECIATION' in line.upper() or 'EXPENSE' in line.upper()):
                        header_idx = i
                        break

                if header_idx == -1:
                    continue

                current_category = ''

                for line in lines[header_idx + 1:]:
                    line = line.strip()
                    if not line:
                        continue

                    if line.upper().startswith('TOTAL') or line.upper().startswith('GRAND TOTAL'):
                        break

                    # Extract amounts from line
                    amounts = re.findall(r'-?[\d,]+\.\d{2}', line)

                    if not amounts:
                        # Might be a category header
                        if not any(c.isdigit() for c in line):
                            current_category = line
                        continue

                    # Get asset name
                    asset_line = line
                    for amt in amounts:
                        asset_line = asset_line.replace(amt, '', 1)
                    asset_name = asset_line.strip()

                    if not asset_name:
                        continue

                    beginning = self.parse_amount(amounts[0]) if len(amounts) >= 1 else 0.0
                    expense = self.parse_amount(amounts[1]) if len(amounts) >= 2 else 0.0
                    ending = self.parse_amount(amounts[2]) if len(amounts) >= 3 else 0.0

                    results.append({
                        "assetName": asset_name,
                        "assetId": None,
                        "period": "",
                        "beginningBalance": beginning,
                        "depreciationExpense": expense,
                        "endingBalance": ending,
                        "category": current_category,
                        "depreciationMethod": "",
                    })

        print(f"[DEPRECIATION-PARSER] Parsed {len(results)} depreciation lines from PDF")
        return results


def main():
    parser = argparse.ArgumentParser(description='Convert Depreciation Schedule reports to JSON format')
    parser.add_argument('input', help='Input file (CSV, XLSX, or PDF)')
    parser.add_argument('-o', '--output', help='Output JSON file (default: print to stdout)')

    args = parser.parse_args()

    converter = DepreciationScheduleConverter()
    input_path = Path(args.input)

    if not input_path.exists():
        print(f"Error: {input_path} does not exist", file=sys.stderr)
        sys.exit(1)

    try:
        if args.output:
            result = converter.convert_to_json(input_path, Path(args.output))
            print(result)
        else:
            print(converter.convert_to_json(input_path))
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
