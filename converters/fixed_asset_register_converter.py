#!/usr/bin/env python3
"""
Fixed Asset Register Converter
Converts CSV, XLSX, and PDF fixed asset register/listing reports to JSON array format.

Output format: flat array of fixed asset records.
"""

import csv
import re
import sys
from pathlib import Path
import argparse
from typing import Dict, List, Any, Optional

from base_converter import BaseConverter, XLSX_SUPPORT, PDF_SUPPORT

if XLSX_SUPPORT:
    import openpyxl
if PDF_SUPPORT:
    import pdfplumber


class FixedAssetRegisterConverter(BaseConverter):
    """Converts Fixed Asset Register reports to JSON array format."""

    def __init__(self, **kwargs):
        super().__init__(use_account_lookup=False)

    def parse_csv(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse CSV fixed asset register."""
        results = []

        with open(filepath, 'r', encoding='utf-8') as f:
            all_lines = f.readlines()

        # Find header row
        header_line_idx = -1
        for idx, line in enumerate(all_lines):
            line_lower = line.lower()
            if 'asset' in line_lower and ('cost' in line_lower or 'acquisition' in line_lower or 'date' in line_lower):
                header_line_idx = idx
                break

        if header_line_idx == -1:
            print("[FIXED-ASSET-PARSER] WARNING: No header row found")
            return results

        from io import StringIO
        csv_content = ''.join(all_lines[header_line_idx:])
        reader = csv.DictReader(StringIO(csv_content))

        print(f"[FIXED-ASSET-PARSER] Column headers: {reader.fieldnames}")

        for row in reader:
            asset_name = (row.get('Asset') or row.get('Asset Name') or
                          row.get('Description') or row.get('ASSET') or '').strip()

            if not asset_name or asset_name.upper() in ('TOTAL', 'GRAND TOTAL'):
                continue

            acquisition_date = (row.get('Acquisition Date') or row.get('Date Acquired') or
                                row.get('Purchase Date') or row.get('Date') or '').strip()
            acquisition_cost = self.parse_amount(
                row.get('Acquisition Cost') or row.get('Cost') or
                row.get('Original Cost') or row.get('Purchase Price') or '0'
            )
            accum_depreciation = self.parse_amount(
                row.get('Accumulated Depreciation') or row.get('Accum. Depreciation') or
                row.get('Total Depreciation') or '0'
            )
            net_book_value = self.parse_amount(
                row.get('Net Book Value') or row.get('NBV') or
                row.get('Book Value') or row.get('Carrying Amount') or '0'
            )
            category = (row.get('Category') or row.get('Asset Class') or
                         row.get('Type') or row.get('Group') or '').strip()
            useful_life = row.get('Useful Life') or row.get('Life (Years)') or ''
            useful_life_years = None
            if useful_life:
                try:
                    useful_life_years = float(str(useful_life).replace('years', '').replace('yrs', '').strip())
                except ValueError:
                    pass

            method = (row.get('Method') or row.get('Depreciation Method') or '').strip()
            disposal_date = (row.get('Disposal Date') or row.get('Date Disposed') or '').strip() or None
            disposal_amount = self.parse_amount(
                row.get('Disposal Amount') or row.get('Proceeds') or row.get('Sale Price') or '0'
            )

            # Determine status
            status = 'Active'
            if disposal_date:
                status = 'Disposed'
            elif net_book_value == 0 and acquisition_cost > 0:
                status = 'Fully Depreciated'

            results.append({
                "assetName": asset_name,
                "assetId": None,
                "category": category,
                "acquisitionDate": acquisition_date,
                "acquisitionCost": acquisition_cost,
                "accumulatedDepreciation": accum_depreciation,
                "netBookValue": net_book_value,
                "usefulLifeYears": useful_life_years,
                "depreciationMethod": method,
                "disposalDate": disposal_date,
                "disposalAmount": disposal_amount,
                "status": status,
            })

        print(f"[FIXED-ASSET-PARSER] Parsed {len(results)} assets")
        return results

    def parse_xlsx(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse XLSX fixed asset register."""
        self.check_xlsx_support()

        results = []

        workbook = openpyxl.load_workbook(filepath, data_only=True)
        sheet = workbook.active

        # Find header row
        header_row = None
        for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row and any(cell and 'asset' in str(cell).lower() for cell in row):
                if any(cell and ('cost' in str(cell).lower() or 'acquisition' in str(cell).lower() or 'date' in str(cell).lower()) for cell in row):
                    header_row = idx
                    break

        if not header_row:
            raise ValueError("Could not find header row in XLSX file")

        headers = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        col_map = {str(header).strip(): idx for idx, header in enumerate(headers) if header}

        # Map columns flexibly
        def find_col(*names):
            for n in names:
                if n in col_map:
                    return col_map[n]
            return None

        asset_col = find_col('Asset', 'Asset Name', 'Description') or 0
        date_col = find_col('Acquisition Date', 'Date Acquired', 'Purchase Date', 'Date')
        cost_col = find_col('Acquisition Cost', 'Cost', 'Original Cost', 'Purchase Price')
        accum_col = find_col('Accumulated Depreciation', 'Accum. Depreciation', 'Total Depreciation')
        nbv_col = find_col('Net Book Value', 'NBV', 'Book Value', 'Carrying Amount')
        cat_col = find_col('Category', 'Asset Class', 'Type', 'Group')
        life_col = find_col('Useful Life', 'Life (Years)')
        method_col = find_col('Method', 'Depreciation Method')
        disp_date_col = find_col('Disposal Date', 'Date Disposed')
        disp_amt_col = find_col('Disposal Amount', 'Proceeds', 'Sale Price')

        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            # The asset register ends at the TOTAL row; anything after it (e.g. a
            # trailing "Depreciation by ..." schedule) is not a fixed asset. Check
            # the first column, since the TOTAL label sits there rather than in the
            # asset-name column.
            first_cell = str(row[0]).strip() if row and row[0] is not None else ''
            if first_cell.upper() in ('TOTAL', 'GRAND TOTAL') or first_cell.lower().startswith('depreciation'):
                break

            if not row or not row[asset_col]:
                continue

            asset_name = str(row[asset_col]).strip()

            acquisition_cost = self.parse_amount(row[cost_col]) if cost_col is not None else 0.0
            accum_depreciation = self.parse_amount(row[accum_col]) if accum_col is not None else 0.0
            net_book_value = self.parse_amount(row[nbv_col]) if nbv_col is not None else 0.0

            useful_life_years = None
            if life_col is not None and row[life_col]:
                try:
                    useful_life_years = float(str(row[life_col]).replace('years', '').replace('yrs', '').strip())
                except ValueError:
                    pass

            disposal_date = str(row[disp_date_col] or '') if disp_date_col is not None else None
            disposal_date = disposal_date if disposal_date else None
            disposal_amount = self.parse_amount(row[disp_amt_col]) if disp_amt_col is not None else 0.0

            status = 'Active'
            if disposal_date:
                status = 'Disposed'
            elif net_book_value == 0 and acquisition_cost > 0:
                status = 'Fully Depreciated'

            results.append({
                "assetName": asset_name,
                "assetId": None,
                "category": str(row[cat_col] or '') if cat_col is not None else '',
                "acquisitionDate": str(row[date_col] or '') if date_col is not None else '',
                "acquisitionCost": acquisition_cost,
                "accumulatedDepreciation": accum_depreciation,
                "netBookValue": net_book_value,
                "usefulLifeYears": useful_life_years,
                "depreciationMethod": str(row[method_col] or '') if method_col is not None else '',
                "disposalDate": disposal_date,
                "disposalAmount": disposal_amount,
                "status": status,
            })

        print(f"[FIXED-ASSET-PARSER] Parsed {len(results)} assets from XLSX")
        return results

    def parse_pdf(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse PDF fixed asset register."""
        self.check_pdf_support()

        results = []

        print("[FIXED-ASSET-PARSER] Starting PDF parse")

        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue

                lines = text.split('\n')

                header_idx = -1
                for i, line in enumerate(lines):
                    if 'ASSET' in line.upper() and ('COST' in line.upper() or 'VALUE' in line.upper()):
                        header_idx = i
                        break

                if header_idx == -1:
                    continue

                current_category = ''

                for line in lines[header_idx + 1:]:
                    line = line.strip()
                    if not line:
                        continue

                    if line.upper().startswith('TOTAL') or line.upper().startswith('GRAND TOTAL'):
                        break

                    amounts = re.findall(r'-?[\d,]+\.\d{2}', line)

                    if not amounts:
                        if not any(c.isdigit() for c in line):
                            current_category = line
                        continue

                    asset_line = line
                    for amt in amounts:
                        asset_line = asset_line.replace(amt, '', 1)
                    asset_name = asset_line.strip()

                    if not asset_name:
                        continue

                    cost = self.parse_amount(amounts[0]) if len(amounts) >= 1 else 0.0
                    accum = self.parse_amount(amounts[1]) if len(amounts) >= 2 else 0.0
                    nbv = self.parse_amount(amounts[2]) if len(amounts) >= 3 else 0.0

                    status = 'Active'
                    if nbv == 0 and cost > 0:
                        status = 'Fully Depreciated'

                    results.append({
                        "assetName": asset_name,
                        "assetId": None,
                        "category": current_category,
                        "acquisitionDate": "",
                        "acquisitionCost": cost,
                        "accumulatedDepreciation": accum,
                        "netBookValue": nbv,
                        "usefulLifeYears": None,
                        "depreciationMethod": "",
                        "disposalDate": None,
                        "disposalAmount": 0.0,
                        "status": status,
                    })

        print(f"[FIXED-ASSET-PARSER] Parsed {len(results)} assets from PDF")
        return results


def main():
    parser = argparse.ArgumentParser(description='Convert Fixed Asset Register reports to JSON format')
    parser.add_argument('input', help='Input file (CSV, XLSX, or PDF)')
    parser.add_argument('-o', '--output', help='Output JSON file (default: print to stdout)')

    args = parser.parse_args()

    converter = FixedAssetRegisterConverter()
    input_path = Path(args.input)

    if not input_path.exists():
        print(f"Error: {input_path} does not exist", file=sys.stderr)
        sys.exit(1)

    try:
        if args.output:
            result = converter.convert_to_json(input_path, Path(args.output))
            print(result)
        else:
            print(converter.convert_to_json(input_path))
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
