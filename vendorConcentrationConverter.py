#!/usr/bin/env python3
"""
Vendor Concentration (Expenses by Vendor Summary) Converter
Converts CSV, XLSX, and PDF Expenses by Vendor Summary reports to simplified JSON array format
"""

import csv
import re
import sys
from pathlib import Path
import argparse
from typing import Dict, List, Any, Optional

from base_converter import BaseConverter, XLSX_SUPPORT, PDF_SUPPORT

if XLSX_SUPPORT:
    import openpyxl
if PDF_SUPPORT:
    import pdfplumber


class VendorConcentrationConverter(BaseConverter):
    """Converts Expenses by Vendor Summary reports to simplified JSON array format"""

    def __init__(self, **kwargs):
        super().__init__(use_account_lookup=False)

    def calculate_percentages(self, vendors: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Calculate percentage of total for each vendor"""
        grand_total = sum(v['payments'] for v in vendors)

        for vendor in vendors:
            if grand_total > 0:
                vendor['percentage'] = (vendor['payments'] / grand_total) * 100
            else:
                vendor['percentage'] = 0.0

        vendors.sort(key=lambda x: x['payments'], reverse=True)
        return vendors

    def parse_csv(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse CSV file and convert to simplified JSON array format"""
        vendors = []

        with open(filepath, 'r', encoding='utf-8') as f:
            for _ in range(4):
                f.readline()

            reader = csv.DictReader(f)
            for row in reader:
                vendor_name = row.get('Vendor', '').strip()

                if vendor_name.upper() == 'TOTAL':
                    break

                if not vendor_name:
                    continue

                total = self.parse_amount(row.get('Total', '0'))

                vendors.append({
                    'vendorName': vendor_name,
                    'payments': total,
                    'percentage': 0.0
                })

        return self.calculate_percentages(vendors)

    def parse_xlsx(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse XLSX file and convert to simplified JSON array format"""
        self.check_xlsx_support()

        vendors = []

        workbook = openpyxl.load_workbook(filepath, data_only=True)
        sheet = workbook.active

        header_row = None
        for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row and any('Vendor' in str(cell) for cell in row if cell):
                header_row = idx
                break

        if not header_row:
            raise ValueError("Could not find header row in XLSX file")

        headers = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        col_map = {str(header).strip(): idx for idx, header in enumerate(headers) if header}

        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[col_map.get('Vendor', 0)]:
                continue

            vendor_name = str(row[col_map.get('Vendor', 0)]).strip()

            if vendor_name.upper() == 'VENDOR':
                continue

            if 'Sandbox Company' in vendor_name or 'Expenses by Vendor' in vendor_name:
                continue

            if any(month in vendor_name for month in ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']):
                continue

            if vendor_name.upper() == 'TOTAL':
                break

            total = self.parse_amount(str(row[col_map.get('Total', 1)] or '0'))

            vendors.append({
                'vendorName': vendor_name,
                'payments': total,
                'percentage': 0.0
            })

        return self.calculate_percentages(vendors)

    def parse_pdf(self, filepath: Path) -> List[Dict[str, Any]]:
        """Parse PDF file and convert to simplified JSON array format"""
        self.check_pdf_support()

        vendors = []

        print("[VENDOR-CONC-PARSER] Starting PDF parse")

        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue

                lines = text.split('\n')
                print(f"[VENDOR-CONC-PARSER] Page has {len(lines)} lines")

                header_idx = -1
                for i, line in enumerate(lines):
                    if 'VENDOR' in line.upper() and 'TOTAL' in line.upper():
                        header_idx = i
                        print(f"[VENDOR-CONC-PARSER] Found header at line {i}")
                        break

                if header_idx == -1:
                    continue

                for line in lines[header_idx + 1:]:
                    line = line.strip()
                    if not line:
                        continue

                    if line.upper().startswith('TOTAL'):
                        print(f"[VENDOR-CONC-PARSER] Found TOTAL line, stopping parse")
                        break

                    amounts = re.findall(r'-?[\d,]+\.\d{2}', line)

                    if not amounts:
                        continue

                    vendor_line = line
                    for amt in amounts:
                        vendor_line = vendor_line.replace(amt, '', 1)
                    vendor_name = vendor_line.strip()

                    if not vendor_name:
                        continue

                    total = self.parse_amount(amounts[-1])

                    vendors.append({
                        'vendorName': vendor_name,
                        'payments': total,
                        'percentage': 0.0
                    })
                    print(f"[VENDOR-CONC-PARSER] Added vendor: {vendor_name} (${total})")

        print(f"[VENDOR-CONC-PARSER] Parsed {len(vendors)} vendors")

        return self.calculate_percentages(vendors)


def main():
    parser = argparse.ArgumentParser(description='Convert Expenses by Vendor Summary reports to JSON format')
    parser.add_argument('input', help='Input file (CSV, XLSX, or PDF)')
    parser.add_argument('-o', '--output', help='Output JSON file (default: print to stdout)')
    parser.add_argument('--batch', action='store_true', help='Process all files in a directory')

    args = parser.parse_args()

    converter = VendorConcentrationConverter()

    if args.batch:
        input_path = Path(args.input)
        if not input_path.is_dir():
            print(f"Error: {input_path} is not a directory", file=sys.stderr)
            sys.exit(1)

        output_dir = Path(args.output) if args.output else input_path.parent / 'converted'
        output_dir.mkdir(exist_ok=True)

        for file in input_path.glob('*'):
            if file.suffix.lower() in ['.csv', '.xlsx', '.pdf']:
                try:
                    output_file = output_dir / f"{file.stem}_vendor_concentration.json"
                    result = converter.convert_to_json(file, output_file)
                    print(result)
                except Exception as e:
                    print(f"Error processing {file}: {e}", file=sys.stderr)
    else:
        input_path = Path(args.input)
        if not input_path.exists():
            print(f"Error: {input_path} does not exist", file=sys.stderr)
            sys.exit(1)

        try:
            if args.output:
                result = converter.convert_to_json(input_path, Path(args.output))
                print(result)
            else:
                print(converter.convert_to_json(input_path))
        except Exception as e:
            print(f"Error: {e}", file=sys.stderr)
            sys.exit(1)


if __name__ == '__main__':
    main()
